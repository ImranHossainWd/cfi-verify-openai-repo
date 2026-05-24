"""
California Fruit Inc — Sorting-Quality Packet Verifier (production v2)

Differences from v1:
  • Sub-packet splitter — multi-WO packets (e.g. Balcorp WO 11555 + 11560 sharing
    one PO) are split by SQR Checkoff List boundaries and verified per sub-packet.
  • Pluggable OCR — Tesseract for printed text, vision-LLM API for handwriting.
  • Customer + spec config in YAML — no code changes to add a customer.
  • Spec validity checks — moisture/sulfur/aflatoxin against per-customer spec.
  • Boilerplate filtering — product extractor ignores SQR Checkoff template text.
  • Numerical reconciliation per sub-packet, not per packet.
  • Marked-up PDF uses pypdf overlay onto the ORIGINAL PDF (preserves quality).
  • CLI takes a PDF directly and runs the full pipeline.

Usage:
    python -m sqr_verifier_v2 verify INPUT.pdf -o OUTPUT_DIR

    # Or programmatically:
    from src.verifier import verify_pdf
    report = verify_pdf("Olive_Nation_packet.pdf", "out/")
"""

from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from PIL import Image, ImageDraw, ImageFont, ImageFile
ImageFile.LOAD_TRUNCATED_IMAGES = True

from ocr_backend import (
    HybridOCR, OCRConfig,
    TesseractBackend, MockVisionBackend, AnthropicVisionBackend,
)

warnings.filterwarnings("ignore")

# =============================================================================
# Config loading
# =============================================================================

@dataclass
class CustomerProfile:
    canonical: str
    aliases: List[str]
    customer_code: str = ""
    co_packer_route: bool = False
    requires_bol: bool = True
    requires_trailer_inspection: bool = True
    is_backup_source_only: bool = False
    notes: str = ""


@dataclass
class SpecRange:
    min: Optional[float] = None
    max: Optional[float] = None
    marginal_tolerance: float = 0.5


@dataclass
class ProductSpec:
    customer: str
    product: str
    moisture_pct: SpecRange = field(default_factory=SpecRange)
    sulfur_ppm: SpecRange = field(default_factory=SpecRange)
    aflatoxin_ppb_max: Optional[float] = None
    total_defect_pct_max: float = 10.0


@dataclass
class Config:
    customers: List[CustomerProfile]
    specs: List[ProductSpec]
    product_aliases: Dict[str, str]
    rules: Dict[str, Any]
    ocr_settings: Dict[str, Any]
    output_settings: Dict[str, Any]

    @classmethod
    def load(cls, config_dir: Path) -> "Config":
        cust_data = yaml.safe_load((config_dir / "customers.yaml").read_text(encoding="utf-8"))
        specs_data = yaml.safe_load((config_dir / "specs.yaml").read_text(encoding="utf-8"))
        rules_data = yaml.safe_load((config_dir / "rules.yaml").read_text(encoding="utf-8"))

        customers = [
            CustomerProfile(
                canonical=c["canonical"],
                aliases=c.get("aliases", []),
                customer_code=c.get("customer_code", ""),
                co_packer_route=c.get("co_packer_route", False),
                requires_bol=c.get("requires_bol", True),
                requires_trailer_inspection=c.get("requires_trailer_inspection", True),
                is_backup_source_only=c.get("is_backup_source_only", False),
                notes=c.get("notes", ""),
            )
            for c in cust_data["customers"]
        ]

        specs = []
        for s in specs_data.get("specs", []):
            mp = s.get("moisture_pct") or {}
            sp = s.get("sulfur_ppm") or {}
            specs.append(ProductSpec(
                customer=s["customer"],
                product=s["product"],
                moisture_pct=SpecRange(
                    min=mp.get("min"), max=mp.get("max"),
                    marginal_tolerance=mp.get("marginal_tolerance", 0.5)),
                sulfur_ppm=SpecRange(min=sp.get("min"), max=sp.get("max")),
                aflatoxin_ppb_max=s.get("aflatoxin_ppb_max"),
                total_defect_pct_max=s.get("total_defect_pct_max", 10.0),
            ))

        product_aliases = specs_data.get("product_aliases", {})
        return cls(customers=customers, specs=specs,
                   product_aliases=product_aliases,
                   rules=rules_data.get("rules", {}),
                   ocr_settings=rules_data.get("ocr", {}),
                   output_settings=rules_data.get("output", {}))

    def find_customer(self, name: str) -> Optional[CustomerProfile]:
        if not name:
            return None
        nl = name.lower().strip()
        for c in self.customers:
            for a in c.aliases:
                if a.lower() == nl or a.lower() in nl or nl in a.lower():
                    return c
        return None

    def find_spec(self, customer_canonical: str, product: str) -> Optional[ProductSpec]:
        # Resolve product code → canonical product name
        canonical_product = self.product_aliases.get(product, product)
        # Exact match first, then "(any)"
        for s in self.specs:
            if s.customer == customer_canonical and s.product == canonical_product:
                return s
        for s in self.specs:
            if s.customer == customer_canonical and s.product == "(any)":
                return s
        return None


# =============================================================================
# Form classification
# =============================================================================

# (priority, keyword_in_upper_text, label, code)
FORM_TAGS = [
    (15, "EXTRA CASE - SORTING QUALITY REPORT", "SQR (Extra Case)", "SQR_XC"),
    (15, "EXTRA CASES USED FROM COLD STORAGE", "Extra Cases USED form", "XC_USED"),
    (14, "SORTING QUALITY REPORT\nCHECKOFF LIST", "SQR Checkoff List", "SQR_CHK"),
    (12, "CHECKOFF LIST", "SQR Checkoff List", "SQR_CHK"),
    (10, "CONTAINER WORKMANSHIP", "Container Workmanship", "CONTAINER"),
    (10, "PRODUCTION REQUEST", "Production Request", "PROD_REQ"),
    (10, "CERTIFICATE OF ANALYSIS", "Certificate of Analysis (COA)", "COA"),
    (10, "FINAL PACKED PRODUCT SHEET", "Final Packed Product Sheet", "FPP"),
    (10, "BILL OF LADING", "Bill of Lading", "BOL"),
    (10, "TRAILER / CARGO INSPECTION", "Trailer / Cargo Inspection", "TRAILER"),
    (10, "TRAILER/CARGO INSPECTION", "Trailer / Cargo Inspection", "TRAILER"),
    (10, "STAMP LOG", "Stamp Log", "STAMP"),
    (10, "PRETEST", "Pretest sheet", "PRETEST"),
    (10, "BIN TAG", "Bin Tag", "BIN_TAG"),
    (10, "PULL TICKET", "Pull Ticket", "PULL"),
    (10, "SORT OUT FORM", "Sort-Out Form", "SORT_OUT"),
    (10, "SORT-OUT FORM", "Sort-Out Form", "SORT_OUT"),
    (10, "LOOSE METAL DETECTOR FINDINGS", "Loose Metal Detector", "LMD"),
    (10, "CASE METAL DETECTOR FINDINGS", "Case Metal Detector", "CMD"),
    (8,  "SORTING QUALITY REPORT", "SQR / Lab Findings", "SQR_FULL"),
    (8,  "INVOICE", "Invoice (Sage)", "INV"),
    (8,  "PURCHASE ORDER", "Customer PO", "PO"),
    (8,  "DAILY TOTAL", "Daily Totals", "DAILY"),
    (5,  "FedEx",  "Shipping label", "SHIP_LABEL"),
    (5,  "FEDEX",  "Shipping label", "SHIP_LABEL"),
    (5,  "UPS",    "Shipping label", "SHIP_LABEL"),
]

# Lines that look like the SQR Checkoff List boilerplate — used to filter products
SQR_BOILERPLATE_PRODUCTS = {
    "RAISINS", "APPLE RINGS", "PITTED PRUNES",   # template "EXTRA CASES of:" line
    "Apple Rings", "Pitted Prunes",
}


def classify_page(text: str, vision_data: Optional[Dict] = None) -> Tuple[str, str]:
    # Vision data wins if it gave us a form_type_guess
    if vision_data and vision_data.get("form_type_guess"):
        guess = vision_data["form_type_guess"].lower()
        # Map vision guess to our codes
        mapping = {
            "invoice": ("Invoice (Sage)", "INV"),
            "customer_po": ("Customer PO", "PO"),
            "production_request": ("Production Request", "PROD_REQ"),
            "coa": ("Certificate of Analysis (COA)", "COA"),
            "fpp": ("Final Packed Product Sheet", "FPP"),
            "bol": ("Bill of Lading", "BOL"),
            "trailer_cargo": ("Trailer / Cargo Inspection", "TRAILER"),
            "sqr_checkoff": ("SQR Checkoff List", "SQR_CHK"),
            "sqr_extra_case": ("SQR (Extra Case)", "SQR_XC"),
            "sqr_full": ("SQR / Lab Findings", "SQR_FULL"),
            "lab_findings": ("Lab Findings", "LAB"),
            "sort_out_findings": ("Sort-Out Findings", "SORT_FIND"),
            "defect_bag_photo": ("Defect Bag Photo", "PHOTO"),
            "container_workmanship": ("Container Workmanship", "CONTAINER"),
            "pretest": ("Pretest sheet", "PRETEST"),
            "bin_tag": ("Bin Tag", "BIN_TAG"),
            "pull_ticket": ("Pull Ticket", "PULL"),
            "sort_out_form": ("Sort-Out Form", "SORT_OUT"),
            "stamp_log": ("Stamp Log", "STAMP"),
            "extra_cases_used": ("Extra Cases USED form", "XC_USED"),
            "loose_metal_detector": ("Loose Metal Detector", "LMD"),
            "case_metal_detector": ("Case Metal Detector", "CMD"),
            "shipping_label": ("Shipping label", "SHIP_LABEL"),
        }
        if guess in mapping:
            return mapping[guess]

    upper = text.upper()
    compact = re.sub(r"[^A-Z0-9]+", "", upper)
    if "STAMPPLOG" in compact or "STAMPLOG" in compact:
        return "Stamp Log", "STAMP"
    best = (0, "(unidentified)", "UNK")
    for prio, kw, lbl, code in FORM_TAGS:
        if kw.upper() in upper and prio > best[0]:
            best = (prio, lbl, code)
    return best[1], best[2]


# =============================================================================
# Field extraction (combines Tesseract regex + vision JSON)
# =============================================================================

RX_INV   = re.compile(r"Invoice\s*number\s*[:\s]*([A-Z]{2,3}\d{5,8})", re.I)
RX_BOL   = re.compile(r"Bill\s*of\s*Lading\s*#?[:\s]*([A-Z]{2,3}\d{5,8})", re.I)
RX_WO_PATTERNS = [
    re.compile(r"W\.?\s*[O0]\.?\s*#?\s*[:\s]*([0-9]{4,6})", re.I),
    re.compile(r"\bW[O0][#\s]*([0-9]{4,6})\b", re.I),
    re.compile(r"Lot\s*#\s*[:\s]*([0-9]{4,6})\b", re.I),
    re.compile(r"Batch\s*[:\s]*([0-9]{4,6})\b", re.I),
    re.compile(r"Work\s*Order[s]?\s*[:\s#]*([0-9]{4,6})\b", re.I),
    re.compile(r"Original\s*WO[#\s]*([0-9]{4,6})\b", re.I),
]
RX_PO_A  = re.compile(r"\bPO[#\s]*[:\s]*([A-Z]?\d{3,8}[-]?\d{0,4})", re.I)
RX_PO_B  = re.compile(r"P\.O\.\s*#?\s*[:\s]*([A-Z]?\d{3,8}[-]?\d{0,4})", re.I)
RX_REF   = re.compile(r"Your\s*reference\s*[:\s]*(\d{3,8})", re.I)
RX_DATE  = re.compile(r"\b(\d{1,2})[/\-](\d{1,2})[/\-](20\d{2})\b")
RX_CASES = re.compile(r"(\d+(?:\.\d+)?)\s*(?:Case|Cs|case)s?\s*@\s*(\d+(?:\.\d+)?)\s*[Ll]bs?", re.I)
RX_MOIST = re.compile(r"(?:Average\s*Moisture|MOISTURE\s*READING|Moisture\s*reading)\s*:?\s*([\d\.]+)\s*%?", re.I)
RX_SO2   = re.compile(r"(?:SULFUR\s*READING|Sulfur\s*Dioxide|S(?:O|0)2[\s\w]*)\s*:?\s*([\d,]+)\s*ppm", re.I)
RX_CROP  = re.compile(r"Crop\s*Year[s]?\s*[:\s]*(20\d{2}\s*-?\s*20?\d{0,2})", re.I)


def normalize_po(po: str) -> str:
    """Strip PO- prefix, all hyphens/whitespace, and leading zeros for comparison.
    '28-9017-2', '289017-2', 'PO-289017-2', '289017 - 2' all normalize to '28901720→' etc.
    Use the digit-only form so hyphenation differences don't cause false flags.
    """
    if not po:
        return ""
    s = re.sub(r"^PO[-_\s]*", "", str(po), flags=re.I)
    # strip surrounding whitespace then squash all internal whitespace+hyphens
    s = re.sub(r"[\s\-_]+", "", s).upper()
    return s.lstrip("0")


def normalize_date(s: str) -> Optional[str]:
    m = RX_DATE.search(s)
    if not m:
        return None
    mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime(y, mo, d).strftime("%Y-%m-%d")
    except ValueError:
        return None


def extract_fields(text: str, code: str, config: Config,
                   vision_data: Optional[Dict] = None) -> Dict[str, Any]:
    """Pull every cross-referenceable field from a page.

    Vision data (if present) takes priority since it's read directly from the
    image including handwriting; Tesseract regex serves as fallback.
    """
    f: Dict[str, Any] = {}

    if vision_data and not vision_data.get("error"):
        if vision_data.get("form_title_text"):
            f["form_title_text"] = vision_data["form_title_text"]
        if vision_data.get("invoice_number"): f["invoice_no"] = vision_data["invoice_number"]
        if vision_data.get("bol_number"):     f["bol_no"]     = vision_data["bol_number"]
        if vision_data.get("wo_numbers"):
            wos = [w for w in vision_data["wo_numbers"] if str(w).isdigit()]
            if wos:
                f["wo_candidates"] = wos
                f["wo"] = wos[0]
        if vision_data.get("po_numbers"):
            pos = [str(p) for p in vision_data["po_numbers"] if str(p).strip()]
            if pos:
                f["po_candidates"] = pos
                f["po"] = pos[0]
        if vision_data.get("customer_name"):
            cust = vision_data["customer_name"].strip()
            cprof = config.find_customer(cust)
            f["customer"] = cprof.canonical if cprof else cust
            if cprof:
                f["customer_profile"] = cprof.canonical
                f["is_backup_source"] = cprof.is_backup_source_only
        if vision_data.get("product_description"):
            prod = vision_data["product_description"].strip()
            if prod and prod not in SQR_BOILERPLATE_PRODUCTS:
                f["product"] = prod
        if vision_data.get("product_code"):
            f["product_code"] = vision_data["product_code"]
            # Resolve to canonical via aliases
            f.setdefault("product",
                         config.product_aliases.get(vision_data["product_code"],
                                                    vision_data["product_code"]))
        if vision_data.get("dates"):
            f["dates"] = [d.get("value") for d in vision_data["dates"] if d.get("value")]
        if vision_data.get("quantities"):
            qs = vision_data["quantities"]
            if qs:
                q = qs[0]
                if q.get("cases"):       f["cases"]     = float(q["cases"])
                if q.get("lbs_per_case"): f["unit_lbs"] = float(q["lbs_per_case"])
                if q.get("total_lbs"):   f["total_lbs"] = float(q["total_lbs"])
                elif "cases" in f and "unit_lbs" in f:
                    f["total_lbs"] = round(f["cases"] * f["unit_lbs"], 2)
        if vision_data.get("moisture_pct") is not None:
            f["moisture_pct"] = float(vision_data["moisture_pct"])
        if vision_data.get("sulfur_ppm") is not None:
            try:
                f["sulfur_ppm"] = float(str(vision_data["sulfur_ppm"]).replace(",", ""))
            except (ValueError, TypeError):
                pass
        if vision_data.get("crop_year"):
            f["crop_year"] = vision_data["crop_year"]
        if vision_data.get("original_crop_years"):
            f["original_crop_years"] = vision_data["original_crop_years"]
        if vision_data.get("original_line"):
            f["original_line"] = vision_data["original_line"]
        if vision_data.get("total_defect_pct") is not None:
            try:
                f["total_defect_pct"] = float(str(vision_data["total_defect_pct"]).replace("%", ""))
            except (ValueError, TypeError):
                pass
        if vision_data.get("initials_present"):
            f["initials_present"] = vision_data["initials_present"]
            f["initials"] = vision_data["initials_present"]   # alias for backwards-compat
        if vision_data.get("checkbox_status"):
            f["checkbox_status"] = vision_data["checkbox_status"]
            f["checkboxes"]      = vision_data["checkbox_status"]
        if vision_data.get("shipping_inspection_items"):
            f["shipping_inspection_items"] = _normalize_inspection_items(
                vision_data["shipping_inspection_items"]
            )
        if vision_data.get("shipment_decision"):
            f["shipment_decision"] = vision_data["shipment_decision"]
        if vision_data.get("comments_corrective_actions"):
            f["comments_corrective_actions"] = vision_data["comments_corrective_actions"]
        if vision_data.get("is_defect_bag_photo"):
            f["is_photo"] = True
            f["defect_bag_label"] = vision_data.get("defect_bag_label")
        if vision_data.get("metal_detector_findings"):
            f["metal_detector_findings"] = vision_data["metal_detector_findings"]
        if vision_data.get("handwritten_corrections"):
            f["corrections"] = vision_data["handwritten_corrections"]
        dynamic_fields = _flatten_dynamic_fields(vision_data.get("all_fields") or {})
        _apply_dynamic_field_aliases(f, dynamic_fields)
        for k, v in dynamic_fields.items():
            f.setdefault(k, v)
        # ---- Auto-pick-up of every scalar field vision OCR returns ----
        # Reserved keys are processed above; everything else is forwarded
        # so the cross-reference matrix can auto-discover new fields without
        # any YAML edit. Add a key to your vision prompt and it shows up.
        RESERVED = {
            "form_type_guess", "form_title_text", "wo_numbers", "po_numbers",
            "invoice_number", "bol_number", "customer_name",
            "product_description", "product_code", "dates", "quantities",
            "moisture_pct", "sulfur_ppm", "aflatoxin", "crop_year",
            "original_crop_years", "original_line", "total_defect_pct",
            "initials_present", "checkbox_status",
            "shipping_inspection_items", "shipment_decision", "comments_corrective_actions",
            "is_defect_bag_photo", "defect_bag_label",
            "metal_detector_findings", "handwritten_corrections",
            "all_fields", "notes", "raw_text", "backend", "confidence_estimate",
            "char_count", "error",
        }
        for k, v in vision_data.items():
            if k in RESERVED:
                continue
            if v is None:
                continue
            if k in f:                           # don't overwrite already-set
                continue
            f[k] = v
        _apply_dynamic_field_aliases(f, {k: v for k, v in f.items() if k not in RESERVED})

    # Always do Tesseract extraction too (as fallback / supplement)
    m = RX_INV.search(text)
    if m and "invoice_no" not in f: f["invoice_no"] = m.group(1).upper().replace(" ", "")
    m = RX_BOL.search(text)
    if m and "bol_no" not in f: f["bol_no"] = m.group(1).upper().replace(" ", "")

    if "wo" not in f:
        wo_hits = []
        for rx in RX_WO_PATTERNS:
            wo_hits += rx.findall(text)
        wo_hits = [w for w in wo_hits if not (w.startswith("20") and len(w) == 4 and 2020 <= int(w) <= 2030)]
        if wo_hits:
            f["wo_candidates"] = list(dict.fromkeys(wo_hits))
            f["wo"] = f["wo_candidates"][0]

    if "po" not in f:
        po_hits = []
        for rx in (RX_PO_A, RX_PO_B):
            po_hits += rx.findall(text)
        po_hits += RX_REF.findall(text)
        po_hits = [p for p in po_hits if len(p.lstrip("0")) >= 3]
        if po_hits:
            seen, uniq = set(), []
            for p in po_hits:
                if p not in seen:
                    seen.add(p); uniq.append(p)
            f["po_candidates"] = uniq
            f["po"] = uniq[0]

    if "customer" not in f:
        for c in config.customers:
            for a in c.aliases:
                if a.lower() in text.lower():
                    f["customer"] = c.canonical
                    f["is_backup_source"] = c.is_backup_source_only
                    break
            if "customer" in f:
                break

    if "product" not in f:
        # Look for known product hints, but EXCLUDE SQR Checkoff boilerplate
        # which lists "RAISINS, APPLE RINGS, PITTED PRUNES" as form examples.
        is_sqr_chk = code == "SQR_CHK"
        for code_alias, canon in config.product_aliases.items():
            if code_alias in text:
                f["product_code"] = code_alias
                f["product"] = canon
                break

    if "dates" not in f:
        ds = []
        for line in text.split("\n"):
            d = normalize_date(line)
            if d:
                ds.append(d)
        if ds:
            f["dates"] = sorted(set(ds))

    if "cases" not in f:
        m = RX_CASES.search(text)
        if m:
            try:
                f["cases"] = float(m.group(1))
                f["unit_lbs"] = float(m.group(2))
                f["total_lbs"] = round(f["cases"] * f["unit_lbs"], 2)
            except ValueError:
                pass

    if "moisture_pct" not in f:
        m = RX_MOIST.search(text)
        if m:
            try: f["moisture_pct"] = float(m.group(1))
            except ValueError: pass

    if "sulfur_ppm" not in f:
        m = RX_SO2.search(text)
        if m:
            try: f["sulfur_ppm"] = float(m.group(1).replace(",", ""))
            except ValueError: pass

    if "crop_year" not in f:
        m = RX_CROP.search(text)
        if m: f["crop_year"] = re.sub(r"\s+", "", m.group(1))

    return f


# =============================================================================
# Page record
# =============================================================================

@dataclass
class PageRecord:
    page_no: int
    image_path: str
    form_label: str
    form_code: str
    fields: Dict[str, Any] = field(default_factory=dict)
    yellow_pct: float = 0.0
    red_pct: float = 0.0
    is_likely_photo: bool = False
    text_len: int = 0
    ocr_backend_used: str = "tesseract"
    confidence_estimate: float = 0.0
    notes: List[str] = field(default_factory=list)


@dataclass
class CheckResult:
    name: str
    status: str  # "pass" | "fail" | "info"
    detail: str
    pages: List[int] = field(default_factory=list)
    sub_packet: Optional[int] = None


@dataclass
class SubPacket:
    index: int
    pages: List[PageRecord]
    primary_wo: Optional[str] = None
    primary_po: Optional[str] = None
    primary_customer: Optional[str] = None
    primary_product: Optional[str] = None
    cases: Optional[float] = None
    unit_lbs: Optional[float] = None
    total_lbs: Optional[float] = None
    checks: List[CheckResult] = field(default_factory=list)
    cross_ref_map: Dict[str, Dict[Any, List[int]]] = field(default_factory=dict)


@dataclass
class PacketReport:
    packet_name: str
    pages: List[PageRecord] = field(default_factory=list)
    sub_packets: List[SubPacket] = field(default_factory=list)
    customer_profile: Optional[CustomerProfile] = None
    packet_level_checks: List[CheckResult] = field(default_factory=list)

    @property
    def all_checks(self) -> List[CheckResult]:
        out = list(self.packet_level_checks)
        for sp in self.sub_packets:
            out.extend(sp.checks)
        return out

    @property
    def n_pass(self): return sum(1 for c in self.all_checks if c.status == "pass")
    @property
    def n_fail(self): return sum(1 for c in self.all_checks if c.status == "fail")
    @property
    def n_info(self): return sum(1 for c in self.all_checks if c.status == "info")
    @property
    def overall(self): return "PASS" if self.n_fail == 0 else "FAIL"


# =============================================================================
# Sub-packet splitter
# =============================================================================

def split_into_sub_packets(pages: List[PageRecord]) -> List[SubPacket]:
    """
    Split pages into sub-packets. Each SQR Checkoff List starts a new sub-packet.

    Pages before the first CHK ("pre-CHK pages" — typically Invoice, Customer PO,
    Production Request, COAs, FPPs, BOL, Trailer/Cargo for the WHOLE order)
    are routed to the sub-packet whose WO# they reference. If a pre-CHK page
    has a detectable WO#, it goes to the matching sub-packet; otherwise it
    goes to sub-packet 1 (the default).

    If no SQR Checkoff List exists at all, the whole packet is one sub-packet.
    """
    chk_indices = [i for i, p in enumerate(pages) if p.form_code == "SQR_CHK"]
    if not chk_indices:
        return [SubPacket(index=0, pages=list(pages))]

    # Initial boundaries: each sub-packet i = [chk_indices[i-1]+1, chk_indices[i+1]]
    # but we'll first build sub-packets from CHK forward, then route pre-CHK pages.
    boundaries: List[Tuple[int, int]] = []
    for i in range(len(chk_indices)):
        start = chk_indices[i]
        end   = chk_indices[i+1] if i + 1 < len(chk_indices) else len(pages)
        boundaries.append((start, end))

    sub_packets = [SubPacket(index=i, pages=list(pages[s:e]))
                   for i, (s, e) in enumerate(boundaries)]

    # Route pre-CHK pages. Each sub-packet has a "tentative" WO# from its CHK page.
    sub_wo: List[Optional[str]] = []
    for sp in sub_packets:
        chk_page = next((p for p in sp.pages if p.form_code == "SQR_CHK"), None)
        sub_wo.append(chk_page.fields.get("wo") if chk_page else None)

    pre_pages = list(pages[:chk_indices[0]])
    for p in pre_pages:
        # Try to match by WO#, else by product, else default to sub-packet 0
        target = 0
        pwo = p.fields.get("wo")
        if pwo:
            for i, w in enumerate(sub_wo):
                if w == pwo:
                    target = i
                    break
        else:
            # fallback: match by product
            pprod = p.fields.get("product")
            if pprod:
                for i, sp in enumerate(sub_packets):
                    chk = next((q for q in sp.pages if q.form_code == "SQR_CHK"), None)
                    if chk and chk.fields.get("product") == pprod:
                        target = i
                        break
        sub_packets[target].pages.insert(0, p)

    # Sort each sub-packet by page_no so the output order makes sense
    for sp in sub_packets:
        sp.pages.sort(key=lambda p: p.page_no)

    return sub_packets


# =============================================================================
# Verification rules
# =============================================================================

WO_REQUIRED_CODES = {"PROD_REQ", "COA", "FPP", "BOL", "TRAILER", "SQR_CHK",
                     "SQR_XC", "XC_USED", "STAMP", "BIN_TAG", "PULL", "SORT_OUT",
                     "CONTAINER", "DAILY", "LMD", "CMD"}

PO_REQUIRED_CODES = {"INV", "PO", "PROD_REQ", "COA", "FPP", "BOL", "TRAILER",
                     "SQR_CHK", "SQR_XC", "XC_USED", "PULL", "SORT_OUT", "CONTAINER"}

REQUIRED_FORMS_FOR_ANY_PACKET = {
    "INV": "Invoice",
    "PROD_REQ": "Production Request",
    "COA": "Certificate of Analysis",
    "FPP": "Final Packed Product Sheet",
    "SQR_CHK": "SQR Checkoff List",
    "STAMP": "Stamp Log",
}

# Of those, which are required PER SUB-PACKET (i.e. per WO)
# vs at the WHOLE PACKET level (per order, shared across WOs)
PER_SUB_PACKET_FORMS = {"SQR_CHK", "COA", "STAMP"}
PER_PACKET_FORMS     = {"INV", "PROD_REQ", "FPP"}

# Forms that legitimately speak for THIS sub-packet's QC values.
# Other forms (backup-source SQR, Sort-Out Form spanning multiple bin
# crop-years, Bin Tags from many different bins) can carry different
# values without being a defect — exclude them from cross-page matching
# and from per-page spec validity checks.
PRIMARY_QC_FORMS = {"COA", "SQR_XC", "PRETEST", "FPP"}


# ----- Field auto-discovery -----------------------------------------------

# Keys that are internal/metadata or non-scalar — should NOT appear as
# cross-reference rows in the matrix.
INTERNAL_FIELD_KEYS = {
    "wo_candidates", "po_candidates",
    "is_backup_source", "is_photo",
    "checkbox_status", "initials_present", "shipping_inspection_items",
    "packaging_information", "case_metal_detector_verification",
    "handwritten_corrections", "corrections",
    "dates", "quantities",
    "defect_bag_label", "metal_detector_findings",
    "moisture_range", "sulfur_range",
    "labeling", "packaging_style", "packaging_count",
    "total_defect_pct_max",
    "customer_profile",
    "_page_confidence", "_field_confidence",
}

# Known fields that always render in this fixed order at the top of the matrix
# (any auto-discovered fields appear below them, alphabetized).
KNOWN_FIELD_ORDER: List[Tuple[str, str]] = [
    ("wo",                    "WO #"),
    ("po",                    "PO #"),
    ("customer",              "Customer"),
    ("product",               "Product"),
    ("product_code",          "Product code"),
    ("cases",                 "Cases"),
    ("unit_lbs",              "Lbs / case"),
    ("total_lbs",             "Total lbs"),
    ("moisture_pct",          "Moisture %"),
    ("sulfur_ppm",            "Sulfur ppm"),
    ("crop_year",             "Crop year"),
    ("original_crop_years",   "Original crop years"),
    ("original_line",         "Original line"),
    ("carrier",               "Carrier"),
    ("invoice_no",            "Invoice #"),
    ("bol_no",                "BOL #"),
    ("shipment_decision",     "Shipment decision"),
    ("comments_corrective_actions", "Comments / corrective actions"),
    ("total_defect_pct",      "Total defect %"),
]


def _humanize_field_name(key: str) -> str:
    """Auto-generate a display label for an unknown field key."""
    s = key.replace("_", " ").strip()
    s = re.sub(r"^field\s+", "", s)
    # Special abbreviations
    s = s.replace("pct", "%").replace("ppm", "ppm").replace("lbs", "lbs")
    s = s.replace(" no", " #").replace("wo", "WO").replace("po", "PO").replace("bol", "BOL")
    # Title-case but preserve abbreviations
    parts = s.split()
    out = []
    for w in parts:
        if w.upper() in ("WO", "PO", "BOL", "QC", "OK", "NA", "%", "PPM", "ID", "USDA", "FDA", "SQF"):
            out.append(w.upper())
        elif w == "%" or w == "#" or w == "ppm" or w == "lbs":
            out.append(w)
        else:
            out.append(w.capitalize())
    return " ".join(out)


def _dynamic_field_key(label: str) -> str:
    """Stable key for packet-specific fields discovered by vision OCR."""
    s = str(label or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return f"field_{s[:80]}" if s else ""


def _scalarize_dynamic_value(value: Any) -> Optional[Any]:
    if value is None or value == "" or value == [] or value == {}:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, dict)):
        try:
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        except TypeError:
            return str(value)
    return str(value)


def _flatten_dynamic_fields(value: Any, prefix: str = "") -> Dict[str, Any]:
    """Flatten vision `all_fields` into first-class packet matrix rows."""
    out: Dict[str, Any] = {}
    if isinstance(value, dict):
        for raw_key, raw_value in value.items():
            label = f"{prefix} {raw_key}".strip()
            if isinstance(raw_value, dict):
                out.update(_flatten_dynamic_fields(raw_value, label))
            else:
                field_key = _dynamic_field_key(label)
                scalar = _scalarize_dynamic_value(raw_value)
                if field_key and scalar is not None:
                    out[field_key] = scalar
    elif prefix:
        field_key = _dynamic_field_key(prefix)
        scalar = _scalarize_dynamic_value(value)
        if field_key and scalar is not None:
            out[field_key] = scalar
    return out


def _number_from_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"[-+]?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(m.group(0)) if m else None


def _json_or_original(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not ((s.startswith("[") and s.endswith("]")) or (s.startswith("{") and s.endswith("}"))):
        return value
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return value


def _normalize_inspection_items(value: Any) -> Any:
    items = _json_or_original(value)
    if not isinstance(items, list):
        return items
    normalized = []
    for item in items:
        if not isinstance(item, dict):
            normalized.append({"item": str(item), "status": "Unreadable"})
            continue
        out = dict(item)
        status = str(out.get("status") or out.get("result") or out.get("value") or "").strip().lower()
        if status in {"ok", "okay", "pass", "passed", "yes", "y", "checked", "true"}:
            out["status"] = "Okay"
        elif status in {"fail", "failed", "no", "n", "false"}:
            out["status"] = "Fail"
        elif not status:
            out["status"] = "Blank"
        else:
            out["status"] = out.get("status") or status.title()
        normalized.append(out)
    return normalized


def _apply_dynamic_field_aliases(fields: Dict[str, Any], dynamic_fields: Dict[str, Any]) -> None:
    """Promote packet-specific all_fields keys into canonical fields used by rules."""
    for key, value in dynamic_fields.items():
        low = key.lower()
        parsed = _json_or_original(value)
        if "po" not in fields and (
            re.search(r"(^|_)po($|_)", low) or "purchase_order" in low or "purchase order" in low
        ):
            match = re.search(r"\b(?:PO[#:\s-]*)?([A-Za-z]*\d[\w-]{2,})\b", str(value), flags=re.I)
            if match:
                fields["po"] = match.group(1)
        if "total_defect_pct" not in fields and "defect" in low and (
            "pct" in low or "percent" in low or "avg" in low or "average" in low or "%" in str(value)
        ):
            number = _number_from_value(value)
            if number is not None:
                fields["total_defect_pct"] = number
        if "original_crop_years" not in fields and "original" in low and "crop" in low:
            fields["original_crop_years"] = parsed
        if "original_line" not in fields and "original" in low and "line" in low:
            fields["original_line"] = parsed
        if "carrier" not in fields and ("carrier" in low or "trucking" in low):
            fields["carrier"] = parsed
        if "shipping_inspection_items" not in fields and "inspection" in low and (
            "item" in low or "table" in low or "shipping" in low or "trailer" in low or "cargo" in low
        ):
            normalized = _normalize_inspection_items(parsed)
            if isinstance(normalized, list):
                fields["shipping_inspection_items"] = normalized
        if "shipment_decision" not in fields and "shipment" in low and "decision" in low:
            fields["shipment_decision"] = parsed


def discover_field_rows(pages: List[PageRecord],
                        config_known: Optional[List[Tuple[str, str]]] = None
                        ) -> List[Tuple[str, str]]:
    """
    Build the ordered list of (field_key, display_label) for every
    cross-referenceable scalar field that appears in the packet.

    Known fields render first in their canonical order; any field returned
    by vision OCR (or set anywhere on a PageRecord) that we haven't seen
    before is auto-added below them alphabetically. No YAML edits required
    when packets bring new unique fields.
    """
    known = list(config_known or KNOWN_FIELD_ORDER)
    known_keys = {k for k, _ in known}

    discovered: Dict[str, int] = {}
    for p in pages:
        for k, v in (p.fields or {}).items():
            if k in INTERNAL_FIELD_KEYS or k in known_keys:
                continue
            # Only scalar values are cross-referenceable
            if v is None or v == "" or v == [] or v == {}:
                continue
            if not isinstance(v, (str, int, float, bool)):
                continue
            discovered[k] = discovered.get(k, 0) + 1

    # Sort discovered fields by frequency desc, then alphabetic
    auto_rows = sorted(discovered.items(), key=lambda kv: (-kv[1], kv[0]))
    return known + [(k, _humanize_field_name(k)) for k, _ in auto_rows]


def build_cross_ref_map(sp: SubPacket,
                         field_rows: Optional[List[Tuple[str, str]]] = None
                         ) -> Dict[str, Dict[Any, List[int]]]:
    """For each cross-referenceable field, build:
        { field_key: { value: [pages where this value appears] } }
    Excludes backup-source pages so they don't pollute the cross-reference.
    If `field_rows` is None, auto-discovers fields from the pages (so any
    new field returned by vision OCR is picked up automatically)."""
    if field_rows is None:
        field_rows = discover_field_rows(sp.pages)
    fields_to_track = [k for k, _ in field_rows]
    out: Dict[str, Dict[Any, List[int]]] = {k: {} for k in fields_to_track}
    for p in sp.pages:
        if p.fields.get("is_backup_source"):
            continue
        for k in fields_to_track:
            v = p.fields.get(k)
            if v is None:
                continue
            # Normalize PO for matching
            key = normalize_po(v) if k == "po" else v
            try:
                hashable = key if isinstance(key, (str, int, float, bool)) else str(key)
            except Exception:
                hashable = str(key)
            out[k].setdefault(hashable, []).append(p.page_no)
    # Also track dates by label
    for p in sp.pages:
        if p.fields.get("is_backup_source"):
            continue
        for d in p.fields.get("dates", []) or []:
            if isinstance(d, dict):
                lbl = d.get("label", "date")
                val = d.get("value")
            else:
                lbl, val = "date", d
            if not val:
                continue
            key = f"date::{lbl}"
            out.setdefault(key, {}).setdefault(val, []).append(p.page_no)
    return out


def cross_ref_summary(crm: Dict[str, Dict[Any, List[int]]],
                       field: str, on_page: int) -> str:
    """Given a field and the current page, return a short summary like
       'matches pages [3, 4, 5]' or 'only page (no other page has this field)'."""
    if field not in crm:
        return ""
    val_to_pages = crm[field]
    # Find the value present on `on_page`
    cur_value = None
    for v, pgs in val_to_pages.items():
        if on_page in pgs:
            cur_value = v
            break
    if cur_value is None:
        return ""
    other_pages = [p for p in val_to_pages[cur_value] if p != on_page]
    if not other_pages:
        return "only page in packet with this field"
    return f"matches pages {other_pages}"


def determine_primary_values(sp: SubPacket) -> None:
    wo_counter = Counter(); po_counter = Counter()
    cust_counter = Counter(); prod_counter = Counter()
    cases_counter = Counter()
    for p in sp.pages:
        if p.fields.get("is_backup_source"):
            continue                                # ignore CraftMark backup pages
        if "wo" in p.fields:    wo_counter[p.fields["wo"]] += 1
        if "po" in p.fields:    po_counter[normalize_po(p.fields["po"])] += 1
        if "customer" in p.fields: cust_counter[p.fields["customer"]] += 1
        if "product" in p.fields:  prod_counter[p.fields["product"]] += 1
        if "cases" in p.fields:    cases_counter[round(p.fields["cases"])] += 1

    if wo_counter:   sp.primary_wo = wo_counter.most_common(1)[0][0]
    if po_counter:   sp.primary_po = po_counter.most_common(1)[0][0]
    if cust_counter: sp.primary_customer = cust_counter.most_common(1)[0][0]
    if prod_counter: sp.primary_product  = prod_counter.most_common(1)[0][0]
    if cases_counter: sp.cases           = cases_counter.most_common(1)[0][0]


def _is_no_new_wo_xc_used(page: PageRecord) -> bool:
    if page.form_code != "XC_USED":
        return False
    boxes = page.fields.get("checkbox_status") or page.fields.get("checkboxes") or []
    for box in boxes:
        item = str(box.get("item", "")).upper()
        if "NO NEW WO" in item and box.get("checked"):
            return True
    all_fields = page.fields.get("all_fields") or {}
    return bool(all_fields.get("no_new_wo_checkbox_marked"))


def reconcile_no_new_wo_pages(sp: SubPacket) -> None:
    """For Extra Cases USED forms, a checked "NO NEW WO#" means inherit identity."""
    for p in sp.pages:
        if not _is_no_new_wo_xc_used(p):
            continue
        changed = []
        for key, value in (
            ("wo", sp.primary_wo),
            ("po", sp.primary_po),
            ("customer", sp.primary_customer),
            ("product", sp.primary_product),
        ):
            if value and p.fields.get(key) != value:
                old = p.fields.get(key)
                p.fields[key] = value
                changed.append(f"{key}: {old!r} -> {value!r}")
        if changed:
            p.fields["is_backup_source"] = False
            p.notes.append(
                "NO NEW WO# checked on Extra Cases USED form; inherited "
                "sub-packet identity (" + "; ".join(changed) + ")"
            )


def run_subpacket_checks(sp: SubPacket, config: Config,
                          customer_profile: Optional[CustomerProfile]) -> None:
    rules_cfg = config.rules

    # Build the cross-reference map FIRST so every check can cite the pages
    # where the value also appears.
    crm = build_cross_ref_map(sp)
    sp.cross_ref_map = crm  # stash for later (Excel matrix output)

    # 1. Identity match — WO# / PO# / Customer / Product
    if rules_cfg.get("identity_match", {}).get("enabled", True):
        # Build the SET of WOs known for this sub-packet. For multi-WO sub-
        # packets (Pedrick: 1 SQR Checkoff covers 4 WOs), several WOs are
        # legitimate. A page's WO is OK if it's in this set.
        # Backup-source pages are excluded — they reference other-customer WOs.
        known_wos = set()
        for p in sp.pages:
            if p.fields.get("is_backup_source"):
                continue
            if p.fields.get("wo"):
                known_wos.add(p.fields["wo"])
            for w in (p.fields.get("wo_candidates") or []):
                known_wos.add(w)

        # WO# check
        if not sp.primary_wo:
            sp.checks.append(CheckResult(
                "WO# present", "fail", "No WO# detected anywhere in this sub-packet", [],
                sub_packet=sp.index))
        else:
            for p in sp.pages:
                if p.form_code not in WO_REQUIRED_CODES:
                    continue
                # Backup-source pages reference other customers' WOs by design
                if p.fields.get("is_backup_source"):
                    sp.checks.append(CheckResult(
                        f"WO# on {p.form_label} (p{p.page_no})",
                        "info",
                        f"Backup-source page — references original-run WO "
                        f"({p.fields.get('wo', '?')}), not the current order's WO. "
                        f"Not a defect.",
                        [p.page_no], sub_packet=sp.index))
                    continue
                pwo = p.fields.get("wo")
                if pwo is None:
                    # Fallback: search raw text for ANY of the known WOs
                    txt_path = Path(p.image_path).parent / "txt" / f"p-{p.page_no:02d}.txt"
                    raw = txt_path.read_text() if txt_path.exists() else ""
                    for kw in known_wos:
                        if re.search(rf"\b{kw}\b", raw):
                            pwo = kw
                            p.fields["wo"] = kw
                            p.notes.append(f"WO# {kw} matched via fallback string-search")
                            break
                if pwo is None:
                    sp.checks.append(CheckResult(
                        f"WO# on {p.form_label} (p{p.page_no})",
                        "info",
                        f"OCR couldn't read WO# (handwritten/low-quality scan). "
                        f"Vision OCR or human review recommended. "
                        f"Sub-packet WOs: {sorted(known_wos)}.",
                        [p.page_no], sub_packet=sp.index))
                elif pwo in known_wos:
                    src = cross_ref_summary(crm, "wo", p.page_no)
                    if len(known_wos) > 1:
                        msg = f"WO# {pwo} ✓ — one of sub-packet WOs {sorted(known_wos)}"
                    else:
                        msg = f"WO# {pwo} ✓"
                    if src:
                        msg = msg + f" — {src}"
                    sp.checks.append(CheckResult(
                        f"WO# on {p.form_label} (p{p.page_no})",
                        "pass", msg,
                        [p.page_no], sub_packet=sp.index))
                else:
                    # Special case: Extra Cases USED + XC_USED + Pull Ticket
                    # legitimately reference ORIGINAL-source WOs (the bins
                    # that were pulled from). Not a defect.
                    if p.form_code in {"XC_USED", "PULL"}:
                        sp.checks.append(CheckResult(
                            f"WO# on {p.form_label} (p{p.page_no})",
                            "info",
                            f"WO# {pwo} references an original-source WO "
                            f"(bins pulled from), not the current order's WO. "
                            f"Sub-packet WOs: {sorted(known_wos)}.",
                            [p.page_no], sub_packet=sp.index))
                    else:
                        sp.checks.append(CheckResult(
                            f"WO# on {p.form_label} (p{p.page_no})",
                            "fail",
                            f"Page WO# {pwo} not in sub-packet WO set "
                            f"{sorted(known_wos)}",
                            [p.page_no], sub_packet=sp.index))

        # PO# check (with normalization)
        if sp.primary_po:
            for p in sp.pages:
                if p.form_code in PO_REQUIRED_CODES:
                    ppo = p.fields.get("po")
                    if ppo is None:
                        sp.checks.append(CheckResult(
                            f"PO# on {p.form_label} (p{p.page_no})",
                            "info",
                            f"PO# not detected on this page (please confirm visually)",
                            [p.page_no], sub_packet=sp.index))
                        continue
                    if normalize_po(ppo) == sp.primary_po:
                        src = cross_ref_summary(crm, "po", p.page_no)
                        msg = f"PO# {ppo} ✓" + (f" — {src}" if src else "")
                        sp.checks.append(CheckResult(
                            f"PO# on {p.form_label} (p{p.page_no})",
                            "pass", msg,
                            [p.page_no], sub_packet=sp.index))
                    else:
                        # Loose match: substring (handles 289017 vs 289017-2)
                        a = normalize_po(ppo)
                        b = sp.primary_po
                        if a in b or b in a:
                            sp.checks.append(CheckResult(
                                f"PO# on {p.form_label} (p{p.page_no})",
                                "info",
                                f"PO# {ppo} is a partial/truncated read of "
                                f"primary {sp.primary_po} (likely OCR truncation)",
                                [p.page_no], sub_packet=sp.index))
                        else:
                            sp.checks.append(CheckResult(
                                f"PO# on {p.form_label} (p{p.page_no})",
                                "fail",
                                f"Page PO# {ppo} ≠ sub-packet primary {sp.primary_po}",
                                [p.page_no], sub_packet=sp.index))

        # Customer
        if sp.primary_customer:
            for p in sp.pages:
                pc = p.fields.get("customer")
                if pc and pc == sp.primary_customer:
                    src = cross_ref_summary(crm, "customer", p.page_no)
                    sp.checks.append(CheckResult(
                        f"Customer on {p.form_label} (p{p.page_no})",
                        "pass",
                        f"Customer = {pc} ✓" + (f" — {src}" if src else ""),
                        [p.page_no], sub_packet=sp.index))
                elif pc and pc != sp.primary_customer:
                    if p.fields.get("is_backup_source"):
                        sp.checks.append(CheckResult(
                            f"Customer on {p.form_label} (p{p.page_no})",
                            "info",
                            f"{pc} — backup source for extra-case order, not an error",
                            [p.page_no], sub_packet=sp.index))
                    else:
                        sp.checks.append(CheckResult(
                            f"Customer on {p.form_label} (p{p.page_no})",
                            "fail",
                            f"Page customer {pc} ≠ sub-packet primary {sp.primary_customer}",
                            [p.page_no], sub_packet=sp.index))

        # Product (new check — was missing from per-page list)
        if sp.primary_product:
            for p in sp.pages:
                pp = p.fields.get("product")
                if pp and pp == sp.primary_product:
                    src = cross_ref_summary(crm, "product", p.page_no)
                    sp.checks.append(CheckResult(
                        f"Product on {p.form_label} (p{p.page_no})",
                        "pass",
                        f"Product = {pp} ✓" + (f" — {src}" if src else ""),
                        [p.page_no], sub_packet=sp.index))

    # 2. Required forms — only check PER-SUB-PACKET forms here
    # (PER_PACKET forms like Invoice/PR/FPP/BOL/Trailer are checked at packet
    # level by the caller — they're shared across all sub-packets in a multi-WO order.)
    if rules_cfg.get("required_forms", {}).get("enabled", True):
        has = {p.form_code for p in sp.pages}
        for code, name in REQUIRED_FORMS_FOR_ANY_PACKET.items():
            if code not in PER_SUB_PACKET_FORMS:
                continue
            if code in has:
                sp.checks.append(CheckResult(
                    f"Required form: {name}", "pass",
                    "Found in sub-packet", [], sub_packet=sp.index))
            else:
                sp.checks.append(CheckResult(
                    f"Required form: {name}", "fail",
                    "Not detected in this sub-packet", [], sub_packet=sp.index))

    # (Trader Joe's exception + BOL/Trailer-Cargo presence are checked at
    # packet level since they apply to the whole order, not per WO.)

    # 4. Numerical reconciliation — case count consistency.
    # Skip FPP since it's order-level (sums across all WOs in the order).
    #
    # When a sub-packet has MULTIPLE WOs (e.g. Pedrick — one SQR Checkoff
    # covers 4 WOs, each with its own case count), the "primary case count"
    # for the sub-packet isn't a single number. Instead we group case counts
    # BY WO# detected on each page, and verify that pages with the same WO#
    # agree. A page whose case count matches ANY WO's case count is legit.
    if rules_cfg.get("numerical_reconciliation", {}).get("enabled", True):
        # Order-level forms list the SUM of case counts across all WOs in the
        # order (e.g. Trailer/Cargo p11 = 50+30+30+30 = 140 cs for Pedrick).
        # Per-WO forms (COA, SQR-XC, Stamp Log, etc.) list their own WO's count.
        ORDER_LEVEL_FORMS = {"FPP", "BOL", "TRAILER", "INV"}

        # Collect case counts per WO across the sub-packet
        cases_by_wo: Dict[str, Counter] = {}
        for p in sp.pages:
            wo = p.fields.get("wo")
            cs = p.fields.get("cases")
            if wo and cs is not None and p.form_code not in ORDER_LEVEL_FORMS:
                cases_by_wo.setdefault(wo, Counter())[round(cs)] += 1
        # Primary case count per WO = most common
        primary_cases_per_wo = {
            wo: counts.most_common(1)[0][0]
            for wo, counts in cases_by_wo.items() if counts
        }
        # All known case counts in this sub-packet (any WO)
        all_known_cases = {v for v in primary_cases_per_wo.values()}
        # Order-level total = sum of WO-level case counts
        order_total_cases = sum(primary_cases_per_wo.values()) if primary_cases_per_wo else None

        for p in sp.pages:
            if "cases" not in p.fields:
                continue
            page_cases = round(p.fields["cases"])
            page_wo    = p.fields.get("wo")

            # Order-level forms: case count should be the SUM of WO counts
            if p.form_code in ORDER_LEVEL_FORMS:
                if order_total_cases is None:
                    sp.checks.append(CheckResult(
                        f"Case count on {p.form_label} (p{p.page_no})",
                        "info",
                        f"Order-level form ({page_cases} cs); no per-WO totals "
                        f"detected to validate the sum against.",
                        [p.page_no], sub_packet=sp.index))
                elif page_cases == order_total_cases:
                    sp.checks.append(CheckResult(
                        f"Case count on {p.form_label} (p{p.page_no})",
                        "pass",
                        f"{page_cases} cs ✓ — equals order-level sum of WO totals "
                        f"({' + '.join(str(c) for c in primary_cases_per_wo.values())} = {order_total_cases})",
                        [p.page_no], sub_packet=sp.index))
                else:
                    sp.checks.append(CheckResult(
                        f"Case count on {p.form_label} (p{p.page_no})",
                        "fail",
                        f"Order-level total {page_cases} cs ≠ sum of WO totals "
                        f"({' + '.join(str(c) for c in primary_cases_per_wo.values())} = {order_total_cases})",
                        [p.page_no], sub_packet=sp.index))
                continue

            # Per-WO forms
            target = primary_cases_per_wo.get(page_wo)
            if target is not None:
                ok = (page_cases == target)
                detail_match = f"matches WO {page_wo} primary {target} cs"
                detail_miss  = f"≠ WO {page_wo} primary {target} cs"
            else:
                ok = page_cases in all_known_cases
                detail_match = f"matches one of the sub-packet WO case counts {sorted(all_known_cases)}"
                detail_miss  = (f"not in sub-packet WO case counts "
                                f"{sorted(all_known_cases)}" if all_known_cases
                                else "no other case counts to compare against")
            if ok:
                src = cross_ref_summary(crm, "cases", p.page_no)
                sp.checks.append(CheckResult(
                    f"Case count on {p.form_label} (p{p.page_no})",
                    "pass",
                    f"{p.fields['cases']} cs ✓ — {detail_match}"
                    + (f"; {src}" if src else ""),
                    [p.page_no], sub_packet=sp.index))
            else:
                sp.checks.append(CheckResult(
                    f"Case count on {p.form_label} (p{p.page_no})",
                    "fail",
                    f"{p.fields['cases']} cs {detail_miss}",
                    [p.page_no], sub_packet=sp.index))

    # 5. Spec validity (moisture / sulfur)
    # Each page is checked against its OWN product's spec — not the sub-
    # packet's "primary" product. For multi-WO sub-packets (Pedrick: 4
    # products in one Checkoff), the primary product is just one of them;
    # Nectarine COA must be validated against Nectarine spec, not Apricot spec.
    if rules_cfg.get("spec_validity", {}).get("enabled", True):
        primary_spec = config.find_spec(sp.primary_customer or "", sp.primary_product or "")
        if primary_spec is None:
            sp.checks.append(CheckResult(
                "Spec validity check",
                "info",
                f"No spec entry for ({sp.primary_customer} / {sp.primary_product}) — "
                f"please add to specs.yaml to enable spec checking",
                [], sub_packet=sp.index))
        for p in sp.pages:
            # Spec validity ONLY runs on primary-QC forms (COA, SQR
            # Extra-Case, FPP, Pretest). Backup-source pages and any page
            # not in the primary-QC list have legitimately different
            # values OR may carry Tesseract-misread values that would
            # produce false-positive failures.
            if p.fields.get("is_backup_source"):
                continue
            if p.form_code not in PRIMARY_QC_FORMS:
                continue
            # Look up the spec for THIS page's customer + product
            page_product  = p.fields.get("product") or sp.primary_product or ""
            page_customer = p.fields.get("customer") or sp.primary_customer or ""
            spec = config.find_spec(page_customer, page_product) or primary_spec
            if spec is None:
                continue
            if True:  # keep indentation continuity for the original block below
                m = p.fields.get("moisture_pct")
                if m is not None and (spec.moisture_pct.min is not None or spec.moisture_pct.max is not None):
                    lo = spec.moisture_pct.min or 0
                    hi = spec.moisture_pct.max or 1e9
                    tol = spec.moisture_pct.marginal_tolerance
                    if lo - tol <= m <= hi + tol:
                        sp.checks.append(CheckResult(
                            f"Moisture spec on p{p.page_no}",
                            "pass" if (lo <= m <= hi) else "info",
                            f"{m}% ∈ [{lo}, {hi}]" + ("" if lo <= m <= hi else f" — within ±{tol} marginal tolerance"),
                            [p.page_no], sub_packet=sp.index))
                    else:
                        sp.checks.append(CheckResult(
                            f"Moisture spec on p{p.page_no}",
                            "fail",
                            f"{m}% outside spec [{lo}, {hi}] (tolerance ±{tol})",
                            [p.page_no], sub_packet=sp.index))
                s = p.fields.get("sulfur_ppm")
                if s is not None and (spec.sulfur_ppm.min is not None or spec.sulfur_ppm.max is not None):
                    lo = spec.sulfur_ppm.min or 0
                    hi = spec.sulfur_ppm.max or 1e9
                    if lo <= s <= hi:
                        sp.checks.append(CheckResult(
                            f"Sulfur spec on p{p.page_no}",
                            "pass", f"{int(s)} ppm ∈ [{int(lo)}, {int(hi)}]",
                            [p.page_no], sub_packet=sp.index))
                    else:
                        sp.checks.append(CheckResult(
                            f"Sulfur spec on p{p.page_no}",
                            "fail", f"{int(s)} ppm outside spec [{int(lo)}, {int(hi)}]",
                            [p.page_no], sub_packet=sp.index))

    # 6. Defect-photo audit (if Lab Findings or SQR_FULL exist)
    if rules_cfg.get("defect_photos", {}).get("enabled", True):
        photo_pages = [p.page_no for p in sp.pages
                       if p.form_code == "PHOTO" or p.fields.get("is_photo")]
        sqr_full = [p.page_no for p in sp.pages
                    if p.form_code in ("SQR_FULL", "LAB", "SQR_XC")]
        if sqr_full:
            sp.checks.append(CheckResult(
                "Defect-photo audit",
                "info" if photo_pages else "info",
                f"{len(photo_pages)} defect bag photo(s) detected. "
                f"Cross-check that every defect listed on findings sheets has a "
                f"matching photo with a 'WO# {sp.primary_wo} <defect>' sticky note.",
                photo_pages + sqr_full, sub_packet=sp.index))

    # 7. Cross-page numerical reconciliation: moisture / sulfur / crop year
    # (these MUST agree on every page that records them — EXCEPT backup-source
    # pages like an attached CraftMark SQR for the original full run that
    # supplied the extra cases. Those have their own moisture/sulfur values
    # taken at different times and should not be compared.)
    # PRIMARY_QC_FORMS now defined as a module-level constant (see top of file).

    def _cross_match_field(field_key: str, label: str, severity: str = "fail",
                            restrict_to: Optional[set] = None):
        # When a sub-packet has multiple WOs (Pedrick: 1 Checkoff covers 4
        # WOs), each WO has its own COA / SQR Extra-Case / etc. with
        # legitimately different moisture / sulfur / crop year. Cross-match
        # GROUPED BY WO# so each WO's pages are checked against each other,
        # not against pages from a different WO in the same sub-packet.
        # Pages without a detected WO# are pooled into a "no-wo" group.
        groups: Dict[str, Dict[Any, List[int]]] = {}
        for p in sp.pages:
            if p.fields.get("is_backup_source"):
                continue
            if restrict_to is not None and p.form_code not in restrict_to:
                continue
            v = p.fields.get(field_key)
            if v is None:
                continue
            page_wo = p.fields.get("wo") or "_no_wo_"
            groups.setdefault(page_wo, {}).setdefault(v, []).append(p.page_no)
        # Run a separate cross-match per WO group
        for wo_key, values in groups.items():
            wo_tag = f" [WO {wo_key}]" if wo_key != "_no_wo_" else ""
            if len(values) <= 1:
                if values:
                    v, pgs = next(iter(values.items()))
                    sp.checks.append(CheckResult(
                        f"{label} cross-page{wo_tag}", "pass",
                        f"{label} = {v} matches across pages {pgs}",
                        pgs, sub_packet=sp.index))
            else:
                # Multiple distinct values found — flag each non-majority value
                sorted_vals = sorted(values.items(), key=lambda kv: -len(kv[1]))
                primary = sorted_vals[0]
                for v, pgs in sorted_vals[1:]:
                    sp.checks.append(CheckResult(
                        f"{label} cross-page{wo_tag}", severity,
                        f"{label} = {v} on pages {pgs} disagrees with {primary[0]} on pages {primary[1]}",
                        pgs + primary[1], sub_packet=sp.index))
                sp.checks.append(CheckResult(
                    f"{label} cross-page (primary){wo_tag}", "pass",
                    f"{label} = {primary[0]} on pages {primary[1]}",
                    primary[1], sub_packet=sp.index))

    _cross_match_field("moisture_pct", "Moisture",  restrict_to=PRIMARY_QC_FORMS)
    _cross_match_field("sulfur_ppm",   "Sulfur ppm", restrict_to=PRIMARY_QC_FORMS)
    _cross_match_field("crop_year",    "Crop year",  restrict_to=PRIMARY_QC_FORMS)
    _cross_match_field("carrier",      "Carrier")    # carrier should match across BOL/Trailer/FedEx label — no restriction

    # 8. Total weight calculation: cases × lbs/case = total_lbs (per page)
    if rules_cfg.get("numerical_reconciliation", {}).get("enabled", True):
        for p in sp.pages:
            cs   = p.fields.get("cases")
            ulb  = p.fields.get("unit_lbs")
            tot  = p.fields.get("total_lbs")
            if cs is not None and ulb is not None:
                expected = round(cs * ulb, 2)
                if tot is None:
                    sp.checks.append(CheckResult(
                        f"Total weight calc on {p.form_label} (p{p.page_no})",
                        "pass",
                        f"{cs} cs × {ulb} lb = {expected} lb (calculated; no Total field on this form)",
                        [p.page_no], sub_packet=sp.index))
                elif abs(tot - expected) <= 0.5:
                    sp.checks.append(CheckResult(
                        f"Total weight calc on {p.form_label} (p{p.page_no})",
                        "pass",
                        f"{cs} cs × {ulb} lb = {tot} lb ✓", [p.page_no],
                        sub_packet=sp.index))
                else:
                    sp.checks.append(CheckResult(
                        f"Total weight calc on {p.form_label} (p{p.page_no})",
                        "fail",
                        f"{cs} cs × {ulb} lb should be {expected} lb but page says {tot} lb",
                        [p.page_no], sub_packet=sp.index))

    # 9. Trailer/Cargo inspection items completeness
    for p in sp.pages:
        if p.form_code != "TRAILER":
            continue
        items = p.fields.get("shipping_inspection_items", []) if isinstance(p.fields, dict) else []
        if not items:
            sp.checks.append(CheckResult(
                f"Trailer/Cargo inspection items (p{p.page_no})",
                "info",
                "Inspection-item table couldn't be read from page (vision OCR recommended)",
                [p.page_no], sub_packet=sp.index))
        else:
            n_ok    = sum(1 for it in items if it.get("status") == "Okay")
            n_fail  = sum(1 for it in items if it.get("status") == "Fail")
            n_blank = len(items) - n_ok - n_fail
            if n_fail > 0:
                # Check each fail has a corrective action
                cca = p.fields.get("comments_corrective_actions", "") or ""
                if cca.strip():
                    sp.checks.append(CheckResult(
                        f"Trailer/Cargo failures with corrective action (p{p.page_no})",
                        "pass",
                        f"{n_fail} fail(s) all have corrective action: {cca[:60]}",
                        [p.page_no], sub_packet=sp.index))
                else:
                    sp.checks.append(CheckResult(
                        f"Trailer/Cargo failures with corrective action (p{p.page_no})",
                        "fail",
                        f"{n_fail} inspection fail(s) but no corrective action recorded",
                        [p.page_no], sub_packet=sp.index))
            elif n_blank > 0:
                sp.checks.append(CheckResult(
                    f"Trailer/Cargo inspection items (p{p.page_no})",
                    "info",
                    f"{n_ok}/{len(items)} items marked Okay; {n_blank} blank/unreadable",
                    [p.page_no], sub_packet=sp.index))
            else:
                sp.checks.append(CheckResult(
                    f"Trailer/Cargo inspection items (p{p.page_no})",
                    "pass",
                    f"All {n_ok}/{len(items)} shipping items marked Okay ✓",
                    [p.page_no], sub_packet=sp.index))
        # Shipment decision
        sd = p.fields.get("shipment_decision") if isinstance(p.fields, dict) else None
        if sd:
            sp.checks.append(CheckResult(
                f"Shipment decision (p{p.page_no})",
                "pass" if sd.lower() == "accepted" else "info",
                f"Marked: {sd}",
                [p.page_no], sub_packet=sp.index))
        # Driver signature
        sigs = p.fields.get("initials_present", []) if isinstance(p.fields, dict) else []
        has_driver = any("driver" in (s.get("location", "").lower()) for s in sigs)
        if has_driver:
            sp.checks.append(CheckResult(
                f"Driver signature (p{p.page_no})",
                "pass", "Driver signature present", [p.page_no],
                sub_packet=sp.index))

    # 10. SQR Checkoff List completeness — all 8 items checked, both signatures
    for p in sp.pages:
        if p.form_code != "SQR_CHK":
            continue
        items = p.fields.get("checkbox_status", []) if isinstance(p.fields, dict) else []
        if items:
            n_checked    = sum(1 for it in items if it.get("checked") is True)
            n_explicitly_unchecked = sum(1 for it in items if it.get("checked") is False)
            if len(items) >= 6:
                sp.checks.append(CheckResult(
                    f"SQR Checkoff items (p{p.page_no})",
                    "pass" if n_checked + n_explicitly_unchecked >= 6 else "info",
                    f"{n_checked} item(s) ticked, {n_explicitly_unchecked} marked NA / NO FINDINGS",
                    [p.page_no], sub_packet=sp.index))
        sigs = p.fields.get("initials_present", []) if isinstance(p.fields, dict) else []
        has_v1 = any("verification" in s.get("location", "").lower() and "2nd" not in s.get("location", "").lower() for s in sigs)
        has_v2 = any("2nd" in s.get("location", "").lower() and "verification" in s.get("location", "").lower() for s in sigs)
        if has_v1:
            sp.checks.append(CheckResult(
                f"SQR Checkoff Verification signature (p{p.page_no})",
                "pass", "Verification signature present", [p.page_no],
                sub_packet=sp.index))
        else:
            sp.checks.append(CheckResult(
                f"SQR Checkoff Verification signature (p{p.page_no})",
                "info", "Verification signature not detected — please confirm",
                [p.page_no], sub_packet=sp.index))
        if has_v2:
            sp.checks.append(CheckResult(
                f"SQR Checkoff 2nd Verification signature (p{p.page_no})",
                "pass", "2nd Verification signature present", [p.page_no],
                sub_packet=sp.index))
        else:
            sp.checks.append(CheckResult(
                f"SQR Checkoff 2nd Verification signature (p{p.page_no})",
                "info", "2nd Verification signature not detected — please confirm",
                [p.page_no], sub_packet=sp.index))

    # 11b. Field-coverage audit — for each page, compare extracted fields
    # against the expected list for that form type. Missing → info note,
    # not a failure (it's a "please review" prompt for the human).
    coverage_cfg = rules_cfg.get("field_coverage_audit", {})
    if coverage_cfg.get("enabled", True):
        expected_map = coverage_cfg.get("expected_fields_by_form", {}) or {}
        for p in sp.pages:
            expected = expected_map.get(p.form_code)
            if not expected:
                continue
            missing = []
            for fkey in expected:
                v = p.fields.get(fkey)
                if v is None or v == "" or v == [] or v == {}:
                    missing.append(fkey)
            if missing:
                sp.checks.append(CheckResult(
                    f"Field coverage on {p.form_label} (p{p.page_no})",
                    "info",
                    f"Expected fields not detected: {', '.join(missing)}. "
                    f"Please verify visually that these fields are absent from the page; "
                    f"if they're present, the page should be re-OCR'd with vision so they're captured.",
                    [p.page_no], sub_packet=sp.index))
            else:
                sp.checks.append(CheckResult(
                    f"Field coverage on {p.form_label} (p{p.page_no})",
                    "pass",
                    f"All {len(expected)} expected fields detected ✓",
                    [p.page_no], sub_packet=sp.index))

    # 11. Defect total within spec (COA pages)
    for p in sp.pages:
        if p.form_code != "COA":
            continue
        d   = p.fields.get("total_defect_pct")
        cap = p.fields.get("total_defect_pct_max", 10.0)
        if d is not None:
            if d <= cap:
                sp.checks.append(CheckResult(
                    f"COA total defect % (p{p.page_no})",
                    "pass", f"{d}% ≤ {cap}% spec ✓",
                    [p.page_no], sub_packet=sp.index))
            else:
                sp.checks.append(CheckResult(
                    f"COA total defect % (p{p.page_no})",
                    "fail", f"{d}% > {cap}% spec — investigate",
                    [p.page_no], sub_packet=sp.index))


# =============================================================================
# Render & OCR pipeline
# =============================================================================

def render_pdf(pdf_path: Path, out_dir: Path, dpi: int = 150) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    def _valid_png(path: Path) -> bool:
        try:
            with Image.open(path) as img:
                img.verify()
            return path.stat().st_size > 0
        except Exception:
            return False

    # PyMuPDF is deterministic in Render's free environment and avoids partial
    # PNGs occasionally produced by timed-out Poppler batch renders.
    try:
        import fitz  # type: ignore
        zoom = dpi / 72
        matrix = fitz.Matrix(zoom, zoom)
        with fitz.open(str(pdf_path)) as doc:
            for idx, page in enumerate(doc, start=1):
                out_path = out_dir / f"p-{idx:02d}.png"
                if _valid_png(out_path):
                    continue
                tmp_path = out_path.with_suffix(".tmp.png")
                pix = page.get_pixmap(matrix=matrix, alpha=False)
                pix.save(str(tmp_path))
                tmp_path.replace(out_path)
        pages = sorted(out_dir.glob("p-*.png"))
        bad = [p for p in pages if not _valid_png(p)]
        if bad:
            raise RuntimeError("Rendered invalid PNG page(s): " + ", ".join(p.name for p in bad))
        return pages
    except ImportError:
        pass

    if not shutil.which("pdftoppm"):
        raise RuntimeError("PDF rendering requires pymupdf or Poppler's pdftoppm.")
    base = out_dir / "p"
    subprocess.run(["pdftoppm", "-r", str(dpi), "-png", str(pdf_path), str(base)],
                   check=True, timeout=180)
    pages = sorted(out_dir.glob("p-*.png"))
    bad = [p for p in pages if not _valid_png(p)]
    if bad:
        raise RuntimeError("Poppler rendered invalid PNG page(s): " + ", ".join(p.name for p in bad))
    return pages


def detect_markings(image_path: str) -> Dict[str, float]:
    import numpy as np
    try:
        im = Image.open(image_path).convert("RGB")
    except Exception:
        return {"yellow_pct": 0.0, "red_pct": 0.0}
    arr = np.array(im)
    h, w = arr.shape[:2]
    R, G, B = arr[:,:,0].astype(int), arr[:,:,1].astype(int), arr[:,:,2].astype(int)
    yellow = (R > 200) & (G > 170) & (B < 160) & ((R - B) > 60)
    red = (R > 140) & (G < 100) & (B < 100) & ((R - G) > 40) & ((R - B) > 40)
    return {
        "yellow_pct": round(yellow.sum() / (h * w) * 100, 3),
        "red_pct":    round(red.sum()    / (h * w) * 100, 3),
    }


def build_pages(image_paths: List[Path], txt_dir: Path, ocr: HybridOCR,
                config: Config) -> List[PageRecord]:
    pages: List[PageRecord] = []
    txt_dir.mkdir(parents=True, exist_ok=True)
    for img in image_paths:
        m = re.match(r"p-(\d+)", img.name)
        if not m: continue
        n = int(m.group(1))
        # Pre-detect markings for OCR escalation decision
        marks = detect_markings(str(img))
        # First-pass OCR with Tesseract — skip if we already have text cached
        cached_txt = txt_dir / f"p-{n:02d}.txt"
        if cached_txt.exists() and cached_txt.stat().st_size > 0:
            text = cached_txt.read_text()
            tess = {"raw_text": text, "char_count": len(text.strip()),
                    "confidence_estimate": 0.5, "backend": "tesseract"}
        else:
            tess = ocr.tess.extract(str(img))
            text = tess.get("raw_text", "")
            cached_txt.write_text(text)
        # Form code from text alone (we'll refine if vision data arrives)
        label, code = classify_page(text)
        meta = {"yellow_pct": marks["yellow_pct"], "red_pct": marks["red_pct"],
                "form_code": code}
        # Possibly escalate to vision OCR
        vision = None
        vision_error = None
        if ocr.should_escalate(tess, meta):
            try:
                vision = ocr.vision.extract(str(img))
            except Exception as e:
                vision_error = str(e)
                vision = {"error": vision_error}
        # Re-classify with vision data if available
        label, code = classify_page(text, vision)
        # Field extraction
        fields = extract_fields(text, code, config, vision_data=vision)
        confidence = float((vision or tess).get("confidence_estimate") or 0.0)
        field_confidence = {
            k: confidence
            for k, v in fields.items()
            if k not in INTERNAL_FIELD_KEYS and v not in (None, "", [], {})
        }
        fields["_page_confidence"] = round(confidence, 3)
        fields["_field_confidence"] = field_confidence
        rec = PageRecord(
            page_no=n,
            image_path=str(img),
            form_label=label,
            form_code=code,
            fields=fields,
            yellow_pct=marks["yellow_pct"],
            red_pct=marks["red_pct"],
            text_len=len(text.strip()),
            ocr_backend_used="vision" if vision and not vision.get("error") else "tesseract",
            confidence_estimate=round(confidence, 3),
        )
        if vision_error:
            rec.notes.append(f"Vision OCR error: {vision_error}")
        if tess.get("error"):
            rec.notes.append(f"Tesseract OCR error: {tess.get('error')}")
        # Photo heuristic
        if (rec.text_len < 80 and (rec.yellow_pct > 0.5 or rec.red_pct > 0.05)) or rec.fields.get("is_photo"):
            rec.is_likely_photo = True
            if not rec.form_code or rec.form_code == "UNK":
                rec.form_label = "Defect Bag Photo"
                rec.form_code = "PHOTO"
        pages.append(rec)
    return pages


# =============================================================================
# PDF annotation
# =============================================================================

AI_PASS_COLOR     = (40, 140, 60)
AI_HIGHLIGHT_COLOR = (255, 240, 130)
AI_FLAG_COLOR     = (235, 130, 25)
AI_INFO_COLOR     = (60, 110, 180)


def annotate_page_image(rec: PageRecord, page_checks: List[CheckResult],
                        out_path: Path) -> None:
    """
    Annotate a page WITHOUT covering any of the original document.

    Layout:
        +--------------------------------------------------+
        |  [TOP BANNER: form type + pass/fail tally]      |  ← new white space
        +--------------------------------------+-----------+
        |                                      |  [INFO    |
        |        ORIGINAL PAGE (unchanged)     |   BOX:    |  ← original document
        |                                      |   fields] |     unmodified
        |                                      |           |
        +--------------------------------------+-----------+
        |  [BOTTOM PANEL: per-page AI checks]              |  ← new white space
        +--------------------------------------------------+

    Banners and panels live in newly-added margins, so nothing on the
    scanned page is ever covered up.
    """
    orig = Image.open(rec.image_path).convert("RGB")
    OW, OH = orig.size

    # New canvas dimensions
    TOP_PAD     = 130    # banner above original
    RIGHT_PAD   = 380    # info box to the right of original
    # Bottom panel sized to fit the per-page check list
    line_count  = max(1, len(page_checks))
    BOTTOM_PAD  = min(900, 60 + line_count * 26 + 30)  # cap at 900px

    NW = OW + RIGHT_PAD
    NH = OH + TOP_PAD + BOTTOM_PAD

    canvas = Image.new("RGB", (NW, NH), "white")
    canvas.paste(orig, (0, TOP_PAD))

    draw = ImageDraw.Draw(canvas, "RGBA")
    try:
        font  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 32)
        med   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
        small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
    except Exception:
        font = med = small = ImageFont.load_default()

    pass_n = sum(1 for c in page_checks if c.status == "pass")
    fail_n = sum(1 for c in page_checks if c.status == "fail")
    info_n = sum(1 for c in page_checks if c.status == "info")

    # ---- TOP BANNER (in new white space) ----
    banner_color = (220, 245, 220, 255) if fail_n == 0 else (255, 220, 200, 255)
    draw.rectangle([(0, 0), (NW, TOP_PAD)], fill=banner_color)
    backend_tag = " [vision OCR]" if rec.ocr_backend_used == "vision" else ""
    draw.text((24, 14),
              f"Page {rec.page_no} — {rec.form_label}{backend_tag}",
              fill=(0, 0, 0), font=font)
    tally = f"AI: {pass_n} matched ✓     {fail_n} flag(s)     {info_n} note(s)"
    tally = f"AI: {pass_n} matched     {fail_n} flag(s)"
    draw.text((24, 70), tally, fill=(0, 0, 0), font=med)

    # Thin separator line
    draw.line([(0, TOP_PAD - 2), (NW, TOP_PAD - 2)], fill=(0, 0, 0), width=2)

    # ---- RIGHT INFO BOX (in new white space alongside original) ----
    field_lines = []
    f = rec.fields
    if f.get("wo"):           field_lines.append(("WO#",      str(f["wo"])))
    if f.get("po"):           field_lines.append(("PO#",      str(f["po"])))
    if f.get("customer"):     field_lines.append(("Cust",     str(f["customer"])[:28]))
    if f.get("product"):      field_lines.append(("Prod",     str(f["product"])[:30]))
    if f.get("cases"):
        field_lines.append(("Cases",    f"{f['cases']} @ {f.get('unit_lbs', '?')} lb"))
    if f.get("moisture_pct"): field_lines.append(("Moisture", f"{f['moisture_pct']}%"))
    if f.get("sulfur_ppm"):   field_lines.append(("SO2",      f"{int(f['sulfur_ppm'])} ppm"))
    if f.get("crop_year"):    field_lines.append(("Crop",     str(f["crop_year"])))
    if f.get("invoice_no"):   field_lines.append(("INV",      str(f["invoice_no"])))
    if f.get("bol_no"):       field_lines.append(("BOL",      str(f["bol_no"])))

    box_x = OW + 16
    box_y = TOP_PAD + 16
    box_w = RIGHT_PAD - 28
    box_h = max(80, 28 * (len(field_lines) + 1) + 16)
    draw.rectangle([(box_x, box_y), (box_x + box_w, box_y + box_h)],
                   fill=(248, 252, 255), outline=AI_INFO_COLOR, width=3)
    draw.text((box_x + 12, box_y + 8), "AI extracted", fill=AI_INFO_COLOR, font=med)
    y = box_y + 40
    for label, value in field_lines:
        draw.text((box_x + 12, y), f"{label}:", fill=(80, 80, 80), font=small)
        draw.text((box_x + 100, y), value, fill=(0, 0, 0), font=small)
        y += 26

    # ---- BOTTOM PANEL (in new white space below original) ----
    panel_y = TOP_PAD + OH + 8
    draw.line([(0, panel_y - 2), (NW, panel_y - 2)], fill=(0, 0, 0), width=2)
    draw.text((24, panel_y + 10), "AI checks on this page:", fill=(0, 0, 0), font=med)
    y = panel_y + 50
    page_checks = [c for c in page_checks if c.status != "info"]
    for c in page_checks[:30]:
        if c.status == "pass":
            mark, color = "✓", AI_PASS_COLOR
        elif c.status == "fail":
            mark, color = "✗", AI_FLAG_COLOR
        else:
            mark, color = "ⓘ", AI_INFO_COLOR
        line = f"{mark} {c.name}: {c.detail}"
        if len(line) > 150:
            line = line[:147] + "..."
        draw.text((40, y), line, fill=color, font=small)
        y += 26

    canvas.save(out_path)


def build_summary_image(report: PacketReport, out_path: Path,
                        page_size=(2100, 2700)) -> None:
    im = Image.new("RGB", page_size, "white")
    draw = ImageDraw.Draw(im)
    W, H = im.size
    try:
        big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 64)
        h1  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 40)
        bd  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
        rg  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 24)
    except Exception:
        big = h1 = bd = rg = ImageFont.load_default()

    banner = (200, 240, 200) if report.overall == "PASS" else (255, 210, 200)
    draw.rectangle([(0, 0), (W, 200)], fill=banner)
    draw.text((40, 30), f"AI VERIFICATION — {report.overall}", fill=(0,0,0), font=big)
    draw.text((40, 110), f"Packet: {report.packet_name}", fill=(0,0,0), font=h1)

    y = 240
    sub_summary = []
    for sp in report.sub_packets:
        sub_summary.append(
            f"#{sp.index+1}: WO {sp.primary_wo or '?'} / PO {sp.primary_po or '?'} / "
            f"{sp.primary_customer or '?'} / {sp.primary_product or '?'}"
        )
    info_rows = [
        ("Sub-packets:", str(len(report.sub_packets))),
        *[(f"  {i+1})", s) for i, s in enumerate(sub_summary)],
        ("Customer:",   report.customer_profile.canonical if report.customer_profile else "(unknown)"),
        ("Total pages:", str(len(report.pages))),
        ("Checks ran:",  str(len(report.all_checks))),
        ("Verified at:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for k, v in info_rows:
        draw.text((60, y), k, fill=(0,0,0), font=bd)
        draw.text((360, y), str(v), fill=(0,0,0), font=rg)
        y += 50

    y += 30
    draw.text((60, y), "Result tally:", fill=(0,0,0), font=h1); y += 60
    draw.text((80, y), f"✓ {report.n_pass} checks passed", fill=AI_PASS_COLOR, font=bd); y += 50
    draw.text((80, y), f"✗ {report.n_fail} flag(s) raised", fill=AI_FLAG_COLOR, font=bd); y += 50
    draw.text((80, y), f"ⓘ {report.n_info} informational note(s)", fill=AI_INFO_COLOR, font=bd); y += 70

    draw.rectangle([(70, y - 70), (900, y - 20)], fill="white")
    y -= 70
    fails = [c for c in report.all_checks if c.status == "fail"]
    if fails:
        draw.text((60, y), "Issues to review:", fill=AI_FLAG_COLOR, font=h1); y += 60
        for i, c in enumerate(fails, 1):
            sp_tag = f" [sub-packet {c.sub_packet+1}]" if c.sub_packet is not None else ""
            line = f"{i}. {c.name}{sp_tag} — {c.detail}"
            if len(line) > 160: line = line[:157] + "..."
            draw.text((80, y), line, fill=(0,0,0), font=rg); y += 36
            if y > H - 220: break
    else:
        draw.text((60, y), "No issues raised. Packet is internally consistent.",
                  fill=AI_PASS_COLOR, font=h1); y += 70

    sig_y = H - 280
    draw.text((60, sig_y),
              "Reviewer signature: _____________________________   Date: ___________",
              fill=(0,0,0), font=bd)
    draw.text((60, sig_y + 60),
              "(Signing here means the AI's findings above have been reviewed and accepted.)",
              fill=(80,80,80), font=rg)
    draw.text((60, H - 100),
              "AI marks: GREEN ✓ matched   ORANGE ✗ flag for review   BLUE ⓘ note   "
              "Pale yellow = AI-reconciled total. Distinct from human reviewer's red/yellow pen.",
              fill=(80,80,80), font=rg)
    draw.rectangle([(50, H - 120), (W - 50, H - 55)], fill="white")
    draw.text((60, H - 100),
              "AI marks: GREEN matched   ORANGE flag for review   "
              "Pale yellow = AI-reconciled total. Distinct from human reviewer's red/yellow pen.",
              fill=(80,80,80), font=rg)
    im.save(out_path)


def assemble_pdf(report: PacketReport, summary_image_path: Path,
                 annotated_dir: Path, out_pdf: Path) -> None:
    pages_imgs = [Image.open(summary_image_path).convert("RGB")]
    for rec in report.pages:
        ann = annotated_dir / f"p-{rec.page_no:02d}_annot.png"
        if ann.exists():
            pages_imgs.append(Image.open(ann).convert("RGB"))
    pages_imgs[0].save(out_pdf, save_all=True, append_images=pages_imgs[1:])


def add_pdf_bookmarks(report: PacketReport, pdf_path: Path) -> None:
    """Add audit-friendly bookmarks to the final generated PDF."""
    try:
        from pypdf import PdfReader, PdfWriter

        reader = PdfReader(str(pdf_path))
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        writer.add_outline_item("Verification Summary", 0)
        for rec in report.pages:
            label = rec.form_label or rec.form_code or f"Page {rec.page_no}"
            writer.add_outline_item(f"p{rec.page_no} - {label}"[:120], rec.page_no)
        tmp = pdf_path.with_suffix(".bookmarked.tmp.pdf")
        with tmp.open("wb") as f:
            writer.write(f)
        tmp.replace(pdf_path)
    except Exception as exc:  # noqa: BLE001
        print(f"  Bookmark generation skipped: {exc}")


# =============================================================================
# Reports (CSV + JSON)
# =============================================================================

def write_issues_csv(report: PacketReport, out_csv: Path) -> None:
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["#", "Status", "Sub-packet", "Check", "Detail",
                    "Pages", "Resolution (human)"])
        idx = 1
        for c in report.all_checks:
            if c.status == "fail":
                sp = (c.sub_packet + 1) if c.sub_packet is not None else ""
                w.writerow([idx, "FLAG", sp, c.name, c.detail,
                            ", ".join(map(str, c.pages)), ""])
                idx += 1


def write_cross_reference_matrix(report: PacketReport, out_xlsx: Path) -> None:
    """
    Excel cross-reference matrix:
      - Rows: every cross-referenceable field (WO#, PO#, Customer, Product, Cases,
              Unit lbs, Total lbs, Moisture, Sulfur, Crop Year, Carrier,
              Invoice #, BOL #, Ship Date, Production Date, Inspection Date, etc.)
      - Columns: every page in the packet (1, 2, 3, ..., N)
      - Cells: the value found on that page for that field, blank if not present
      - Color: green when the value matches the primary, orange when it disagrees
    Plus a 'Field summary' tab listing the canonical value, count of matching
    pages, and any disagreements.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ---------------- Sheet 1: Cross-reference Matrix ----------------
    ws = wb.active
    ws.title = "Cross-reference Matrix"

    HDR_FILL  = PatternFill("solid", start_color="FF1F4E79")
    HDR_FONT  = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    PASS_FILL = PatternFill("solid", start_color="FFD4EFDA")
    INFO_FILL = PatternFill("solid", start_color="FFD9E7F5")
    FAIL_FILL = PatternFill("solid", start_color="FFFFD7BF")
    THIN      = Side(border_style="thin", color="FF888888")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    pages = report.pages
    n_pages = len(pages)

    # Build header
    ws.cell(1, 1, "Field").fill = HDR_FILL
    ws.cell(1, 1).font = HDR_FONT
    ws.cell(1, 2, "Sub-pkt").fill = HDR_FILL
    ws.cell(1, 2).font = HDR_FONT
    ws.cell(1, 3, "Primary value").fill = HDR_FILL
    ws.cell(1, 3).font = HDR_FONT
    for i, p in enumerate(pages):
        c = 4 + i
        ws.cell(1, c, f"p{p.page_no}\n{p.form_label[:18]}").fill = HDR_FILL
        ws.cell(1, c).font = HDR_FONT
        ws.cell(1, c).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 22
    ws.row_dimensions[1].height = 50
    ws.freeze_panes = "D2"

    # Auto-discover all cross-referenceable fields (known + new from vision OCR)
    FIELD_ROWS = discover_field_rows(report.pages)

    # Build a quick lookup: page_no -> {field: value}
    pf = {p.page_no: p.fields for p in pages}

    row = 2
    for sp in report.sub_packets:
        sp_label = f"#{sp.index + 1}"
        crm = sp.cross_ref_map or {}
        sp_pages = {p.page_no for p in sp.pages}
        # Section header for this sub-packet
        ws.cell(row, 1, f"=== Sub-packet {sp_label} (WO {sp.primary_wo or '?'} / PO {sp.primary_po or '?'}) ===")
        ws.cell(row, 1).font = Font(bold=True, italic=True, size=10)
        for c in range(1, 4 + n_pages):
            ws.cell(row, c).fill = PatternFill("solid", start_color="FFEAEAEA")
        row += 1

        for field_key, display in FIELD_ROWS:
            ws.cell(row, 1, display).font = Font(bold=True)
            ws.cell(row, 2, sp_label)
            # Determine "primary" value for this field in this sub-packet
            value_to_pages = crm.get(field_key, {})
            # Filter to only this sub-packet's pages
            value_to_pages_sp = {
                v: [pn for pn in pgs if pn in sp_pages]
                for v, pgs in value_to_pages.items()
            }
            value_to_pages_sp = {v: pgs for v, pgs in value_to_pages_sp.items() if pgs}
            if value_to_pages_sp:
                # primary = most-common
                primary_val = max(value_to_pages_sp.items(), key=lambda kv: len(kv[1]))[0]
                ws.cell(row, 3, str(primary_val))
            else:
                primary_val = None

            # Fill values per page
            for i, p in enumerate(pages):
                c = 4 + i
                if p.page_no not in sp_pages:
                    continue                     # leave blank for other sub-packet
                v = p.fields.get(field_key)
                if v is None:
                    continue
                ws.cell(row, c, str(v))
                ws.cell(row, c).alignment = Alignment(wrap_text=True, horizontal="center")
                ws.cell(row, c).border = BORDER
                # Color
                if primary_val is not None and str(v).strip().lower() == str(primary_val).strip().lower():
                    ws.cell(row, c).fill = PASS_FILL
                else:
                    # Loose match: substring (PO suffix etc.)
                    if primary_val is not None and (str(v) in str(primary_val) or str(primary_val) in str(v)):
                        ws.cell(row, c).fill = INFO_FILL
                    else:
                        ws.cell(row, c).fill = FAIL_FILL
            row += 1

    # ---------------- Sheet 2: Field Summary ----------------
    ws2 = wb.create_sheet("Field Summary")
    headers = ["Sub-pkt", "Field", "Primary value", "# pages with this value",
               "Pages with primary", "Disagreeing values (page list)"]
    for c, h in enumerate(headers, 1):
        ws2.cell(1, c, h).fill = HDR_FILL
        ws2.cell(1, c).font = HDR_FONT
    ws2.row_dimensions[1].height = 24
    widths = [8, 24, 22, 12, 26, 50]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    rr = 2
    for sp in report.sub_packets:
        sp_pages = {p.page_no for p in sp.pages}
        crm = sp.cross_ref_map or {}
        for field_key, display in FIELD_ROWS:
            value_to_pages = crm.get(field_key, {})
            value_to_pages_sp = {
                v: [pn for pn in pgs if pn in sp_pages]
                for v, pgs in value_to_pages.items()
            }
            value_to_pages_sp = {v: pgs for v, pgs in value_to_pages_sp.items() if pgs}
            if not value_to_pages_sp:
                continue
            primary_val, primary_pages = max(value_to_pages_sp.items(), key=lambda kv: len(kv[1]))
            disagreements = [(v, pgs) for v, pgs in value_to_pages_sp.items() if v != primary_val]
            ws2.cell(rr, 1, f"#{sp.index + 1}")
            ws2.cell(rr, 2, display)
            ws2.cell(rr, 3, str(primary_val))
            ws2.cell(rr, 4, len(primary_pages))
            ws2.cell(rr, 5, ", ".join(map(str, sorted(primary_pages))))
            if disagreements:
                disagree_str = "; ".join(f"{v} on pages {sorted(pgs)}" for v, pgs in disagreements)
                ws2.cell(rr, 6, disagree_str)
                ws2.cell(rr, 6).fill = FAIL_FILL
            rr += 1

    # ---------------- Sheet 3: Per-page check trace ----------------
    ws3 = wb.create_sheet("Per-page checks")
    headers = ["Sub-pkt", "Page", "Form", "Status", "Check", "Detail"]
    for c, h in enumerate(headers, 1):
        ws3.cell(1, c, h).fill = HDR_FILL
        ws3.cell(1, c).font = HDR_FONT
    ws3.row_dimensions[1].height = 24
    widths = [8, 8, 24, 10, 36, 80]
    for i, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    rr = 2
    page_form = {p.page_no: p.form_label for p in pages}
    all_checks = [c for c in report.all_checks if c.status != "info"]
    for c in all_checks:
        ws3.cell(rr, 1, f"#{c.sub_packet+1}" if c.sub_packet is not None else "PKT")
        ws3.cell(rr, 2, ", ".join(map(str, c.pages)) if c.pages else "")
        if c.pages:
            ws3.cell(rr, 3, page_form.get(c.pages[0], ""))
        ws3.cell(rr, 4, c.status.upper())
        ws3.cell(rr, 5, c.name)
        ws3.cell(rr, 6, c.detail)
        if c.status == "pass":
            ws3.cell(rr, 4).fill = PASS_FILL
        elif c.status == "fail":
            ws3.cell(rr, 4).fill = FAIL_FILL
        else:
            ws3.cell(rr, 4).fill = INFO_FILL
        rr += 1

    wb.save(out_xlsx)


def write_trace_json(report: PacketReport, out_json: Path) -> None:
    data = {
        "packet_name": report.packet_name,
        "overall": report.overall,
        "tally": {"pass": report.n_pass, "fail": report.n_fail},
        "sub_packets": [
            {
                "index": sp.index,
                "primary_wo": sp.primary_wo,
                "primary_po": sp.primary_po,
                "primary_customer": sp.primary_customer,
                "primary_product": sp.primary_product,
                "page_range": [sp.pages[0].page_no, sp.pages[-1].page_no] if sp.pages else None,
                "cases": sp.cases,
                "checks": [asdict(c) for c in sp.checks if c.status != "info"],
            }
            for sp in report.sub_packets
        ],
        "packet_level_checks": [asdict(c) for c in report.packet_level_checks if c.status != "info"],
        "pages": [
            {
                "page_no": p.page_no,
                "form_label": p.form_label,
                "form_code": p.form_code,
                "ocr_backend_used": p.ocr_backend_used,
                "confidence_estimate": p.confidence_estimate,
                "yellow_pct": p.yellow_pct,
                "red_pct":    p.red_pct,
                "is_photo":   p.is_likely_photo,
                "fields":     p.fields,
            }
            for p in report.pages
        ],
    }
    out_json.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


# =============================================================================
# Top-level driver
# =============================================================================

def verify_pdf(pdf_path: str, out_dir: str,
               config_dir: Optional[str] = None,
               ocr_provider: str = "mock",
               vision_cache_path: Optional[str] = None,
               packet_name: Optional[str] = None) -> PacketReport:
    pdf_p = Path(pdf_path)
    out   = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    cfg_dir = Path(config_dir) if config_dir else Path(__file__).parent.parent / "config"
    config = Config.load(cfg_dir)

    name = packet_name or pdf_p.stem
    work = out / "_work" / name
    work.mkdir(parents=True, exist_ok=True)

    # 1. Render
    print(f"[{name}] Rendering PDF → images...")
    image_paths = render_pdf(pdf_p, work / "pages", dpi=150)
    print(f"  {len(image_paths)} pages rendered.")

    # 2. OCR
    print(f"[{name}] OCR (Tesseract + escalation to {ocr_provider} for handwriting)...")
    ocr_cfg = OCRConfig(
        primary_backend="tesseract",
        handwriting_backend="vision",
        vision_provider=ocr_provider,
        vision_cache_path=vision_cache_path,
        vision_trigger_min_chars=config.ocr_settings.get("vision_trigger_min_chars", 80),
        vision_trigger_marking_pct=config.ocr_settings.get("vision_trigger_marking_pct", 0.5),
        vision_trigger_form_codes=config.ocr_settings.get("vision_trigger_form_codes", []),
        force_vision_all_pages=os.environ.get("FULL_PACKET_FIELD_DISCOVERY", "").strip().lower()
        in {"1", "true", "yes", "on"},
    )
    ocr = HybridOCR(ocr_cfg)
    pages = build_pages(image_paths, work / "txt", ocr, config)
    n_vision = sum(1 for p in pages if p.ocr_backend_used == "vision")
    print(f"  Tesseract: {len(pages)-n_vision} pages, Vision OCR: {n_vision} pages.")

    # 3. Sub-packet split
    print(f"[{name}] Splitting into sub-packets...")
    sub_packets = split_into_sub_packets(pages)
    print(f"  {len(sub_packets)} sub-packet(s) detected.")

    report = PacketReport(packet_name=name, pages=pages, sub_packets=sub_packets)

    # 4. Determine packet customer profile
    cust_counter: Counter = Counter()
    for p in pages:
        if p.fields.get("customer") and not p.fields.get("is_backup_source"):
            cust_counter[p.fields["customer"]] += 1
    if cust_counter:
        primary_cust = cust_counter.most_common(1)[0][0]
        report.customer_profile = config.find_customer(primary_cust)

    # 5. AUTO-DETECT BACKUP-SOURCE PAGES.
    # Any page where the customer differs from the packet's primary customer
    # is a backup-source page — period. Form code doesn't matter; what
    # matters is that the page references somebody else's order (e.g.
    # CraftMark / Trader Joe's original-run SQR attached as backup for
    # an extra-case order). Trader Joe's may be a real customer in their
    # own right elsewhere, but on THIS packet they're a source.
    primary_cust = report.customer_profile.canonical if report.customer_profile else None
    if primary_cust:
        for p in pages:
            page_cust = p.fields.get("customer")
            if (page_cust
                    and page_cust != primary_cust
                    and not _is_no_new_wo_xc_used(p)
                    and not p.fields.get("is_backup_source")):
                p.fields["is_backup_source"] = True
                p.notes.append(
                    f"Auto-flagged backup-source: page customer '{page_cust}' ≠ "
                    f"packet primary '{primary_cust}'"
                )

    # 5b. Per-sub-packet rules
    print(f"[{name}] Running rules per sub-packet...")
    for sp in sub_packets:
        determine_primary_values(sp)
        reconcile_no_new_wo_pages(sp)
        run_subpacket_checks(sp, config, report.customer_profile)

    # 5b. Packet-level rules (forms shared across all WOs in the order)
    has_in_packet = {p.form_code for p in pages}
    for code, fname in REQUIRED_FORMS_FOR_ANY_PACKET.items():
        if code not in PER_PACKET_FORMS:
            continue
        if code in has_in_packet:
            report.packet_level_checks.append(CheckResult(
                f"Order-level form: {fname}", "pass", "Found in packet", []))
        else:
            report.packet_level_checks.append(CheckResult(
                f"Order-level form: {fname}", "fail",
                "Not detected anywhere in packet", []))

    # Trader Joe's / BOL / Trailer-Cargo at packet level
    cprof = report.customer_profile
    if cprof and cprof.co_packer_route:
        report.packet_level_checks.append(CheckResult(
            "Co-packer route", "info",
            f"{cprof.canonical} ships via co-packer; BOL/Trailer-Cargo NOT required",
            []))
    else:
        req_bol = cprof.requires_bol if cprof else True
        req_tc  = cprof.requires_trailer_inspection if cprof else True
        if req_bol:
            if "BOL" in has_in_packet:
                report.packet_level_checks.append(CheckResult(
                    "Bill of Lading present", "pass", "BOL found in packet", []))
            else:
                report.packet_level_checks.append(CheckResult(
                    "Bill of Lading present", "fail", "BOL required but not detected", []))
        if req_tc:
            if "TRAILER" in has_in_packet:
                report.packet_level_checks.append(CheckResult(
                    "Trailer/Cargo Inspection present", "pass", "Found in packet", []))
            else:
                report.packet_level_checks.append(CheckResult(
                    "Trailer/Cargo Inspection present", "fail",
                    "Required but not detected", []))

    # 6. Build outputs
    print(f"[{name}] Annotating pages...")
    page_checks: Dict[int, List[CheckResult]] = defaultdict(list)
    for c in report.all_checks:
        for pg in c.pages:
            page_checks[pg].append(c)
    annotated_dir = out / "annotated_pages"; annotated_dir.mkdir(exist_ok=True)
    for rec in report.pages:
        ann_path = annotated_dir / f"p-{rec.page_no:02d}_annot.png"
        if ann_path.exists() and ann_path.stat().st_size > 0:
            continue            # idempotent — already annotated
        annotate_page_image(rec, page_checks.get(rec.page_no, []), ann_path)
    summary_p = out / f"{name}_summary.png"
    build_summary_image(report, summary_p)
    pdf_out = out / f"{name}_AI_VERIFIED.pdf"
    assemble_pdf(report, summary_p, annotated_dir, pdf_out)
    add_pdf_bookmarks(report, pdf_out)
    write_issues_csv(report, out / f"{name}_issues.csv")
    write_trace_json(report, out / f"{name}_trace.json")
    # Cross-reference matrix is produced as part of the cross_reference_matrix rule
    if config.rules.get("cross_reference_matrix", {}).get("enabled", True):
        matrix_path = out / f"{name}_cross_reference_matrix.xlsx"
        write_cross_reference_matrix(report, matrix_path)
        # Add a meta-check noting that the matrix exists (so it shows up in the
        # PDF summary as evidence of audit-trail compliance)
        report.packet_level_checks.append(CheckResult(
            "Cross-reference matrix",
            "pass",
            f"Excel matrix written ({matrix_path.name}) — full field × page audit trail",
            []))
    print(f"[{name}] Done. Output: {pdf_out}")
    return report


# =============================================================================
# CLI
# =============================================================================

def cli():
    import argparse
    ap = argparse.ArgumentParser(prog="sqr_verifier_v2",
                                 description="Sorting-Quality Packet Verifier (production v2)")
    ap.add_argument("pdf", help="Input scanned packet PDF")
    ap.add_argument("-o", "--out", required=True, help="Output directory")
    ap.add_argument("--name", default=None, help="Packet name (default: PDF stem)")
    ap.add_argument("--config", default=None, help="Config directory (default: ../config)")
    ap.add_argument("--vision-provider", default="mock",
                    choices=["mock", "anthropic", "openai", "google_docai"],
                    help="Vision OCR backend for handwriting pages")
    ap.add_argument("--vision-cache", default=None,
                    help="Path to mock vision cache JSON (for offline runs)")
    args = ap.parse_args()

    rep = verify_pdf(
        pdf_path=args.pdf,
        out_dir=args.out,
        config_dir=args.config,
        ocr_provider=args.vision_provider,
        vision_cache_path=args.vision_cache,
        packet_name=args.name,
    )
    print(f"\n{'='*70}\nPACKET: {rep.packet_name}")
    print(f"OVERALL: {rep.overall}")
    print(f"Sub-packets: {len(rep.sub_packets)}")
    for sp in rep.sub_packets:
        print(f"  Sub-packet #{sp.index+1}: WO={sp.primary_wo} PO={sp.primary_po} "
              f"Cust={sp.primary_customer} Product={sp.primary_product} "
              f"Cases={sp.cases}")
    print(f"Pass: {rep.n_pass}   Fail: {rep.n_fail}   Info: {rep.n_info}")
    if rep.n_fail:
        print("\nFlags:")
        for c in rep.all_checks:
            if c.status == "fail":
                sp_tag = f" [sub#{c.sub_packet+1}]" if c.sub_packet is not None else ""
                print(f"  ✗ {c.name}{sp_tag}: {c.detail}")


if __name__ == "__main__":
    cli()

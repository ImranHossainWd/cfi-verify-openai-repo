from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import threading
import traceback
import urllib.error
import urllib.request
import csv
import hashlib
import hmac
import secrets
import smtplib
import zipfile
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

import yaml
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:
    pass

ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from verifier import verify_pdf  # noqa: E402


APP_NAME = "California Fruit OpenAI Sorting Quality Verifier"
APP_VERSION = "2026-06-10-stamp-log-suppression-v6"
PREVIEW_MAX_ROWS = int(os.environ.get("PREVIEW_MAX_ROWS", "250"))
PREVIEW_MAX_COLS = int(os.environ.get("PREVIEW_MAX_COLS", "60"))
DATA_DIR = Path(os.environ.get("SQR_DATA_DIR", ROOT / "web_data")).resolve()
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
JOBS_FILE = DATA_DIR / "jobs.json"
INSPECTIONS_FILE = DATA_DIR / "inspections.json"
BOARD_FILE = DATA_DIR / "issue_board.json"
TEMPLATES_FILE = DATA_DIR / "form_templates.json"
USERS_FILE = DATA_DIR / "users.json"
SESSIONS_FILE = DATA_DIR / "sessions.json"
FORM_JOBS_FILE = DATA_DIR / "form_jobs.json"
FORM_UPLOAD_DIR = DATA_DIR / "form_uploads"
FORM_OUTPUT_DIR = DATA_DIR / "form_outputs"
TEMPLATE_SAMPLE_DIR = DATA_DIR / "template_samples"
MONTHLY_DIR = DATA_DIR / "monthly"
PRODUCTION_CALENDAR_FILE = DATA_DIR / "production_calendar.json"
LEARNED_RULES_FILE = DATA_DIR / "learned_rules.json"
NOTIFICATIONS_FILE = DATA_DIR / "notifications.json"
BACKUPS_DIR = DATA_DIR / "backups"
FORM_FILING_DIR = DATA_DIR / "filed_forms"
BUNDLED_CONFIG_DIR = ROOT / "config"
CONFIG_DIR = Path(os.environ.get("SQR_CONFIG_DIR", DATA_DIR / "config")).resolve()
VISION_CACHE = Path(os.environ.get("VISION_CACHE_PATH", ROOT / "cache" / "vision_cache.json")).resolve()
DEFAULT_VISION_PROVIDER = os.environ.get("VISION_PROVIDER", "openai").strip().lower() or "openai"
ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-20250514")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5.4-mini")
FULL_PACKET_FIELD_DISCOVERY = os.environ.get("FULL_PACKET_FIELD_DISCOVERY", "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "150"))
MAX_CONCURRENT_PACKET_RUNS = max(1, int(os.environ.get("MAX_CONCURRENT_PACKET_RUNS", "1")))
PROVIDERS = ["openai", "anthropic", "mock"]
RESOLVED_ISSUE_STATUSES = {"Accepted as Is", "False Positive", "Resolved"}
SESSION_COOKIE = "sqr_session"
AUTH_SESSION_DAYS = int(os.environ.get("AUTH_SESSION_DAYS", "14"))
AUTH_COOKIE_SECURE = os.environ.get("AUTH_COOKIE_SECURE", "true").strip().lower() in {"1", "true", "yes", "on"}
ROLE_ORDER = {"Viewer": 1, "Reviewer": 2, "Admin": 3}
CONFIG_EDITABLE_FILES = {
    "rules": "rules.yaml",
    "customers": "customers.yaml",
    "specs": "specs.yaml",
}
DEFAULT_FOOD_SAFETY_TEMPLATES = {
    "AIR_COMPRESSOR": {
        "label": "Air Compressor Inspection",
        "required_terms": ["Air Compressor"],
        "require_date": True,
        "required_signatures": 2,
    },
    "FORKLIFT_DAILY": {
        "label": "Daily Forklift Inspection",
        "required_terms": ["Forklift"],
        "require_date": True,
        "required_signatures": 1,
    },
    "SANITATION": {
        "label": "Sanitation Inspection",
        "required_terms": ["Sanitation"],
        "require_date": True,
        "required_signatures": 1,
    },
    "FIRST_AID": {
        "label": "First Aid Supplies Log",
        "required_terms": ["First Aid"],
        "require_date": True,
        "required_signatures": 1,
    },
}

for directory in (
    DATA_DIR, UPLOAD_DIR, OUTPUT_DIR, CONFIG_DIR, MONTHLY_DIR,
    FORM_UPLOAD_DIR, FORM_OUTPUT_DIR, TEMPLATE_SAMPLE_DIR, BACKUPS_DIR,
    FORM_FILING_DIR,
):
    directory.mkdir(parents=True, exist_ok=True)
for bundled_file in BUNDLED_CONFIG_DIR.glob("*.yaml"):
    runtime_file = CONFIG_DIR / bundled_file.name
    if not runtime_file.exists():
        shutil.copy2(bundled_file, runtime_file)

app = FastAPI(title=APP_NAME)
app.mount("/static", StaticFiles(directory=ROOT / "app" / "static"), name="static")
templates = Jinja2Templates(directory=ROOT / "app" / "templates")
jobs_lock = threading.Lock()
packet_run_semaphore = threading.Semaphore(MAX_CONCURRENT_PACKET_RUNS)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def slugify(value: str) -> str:
    stem = Path(value).stem
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "-", stem).strip(".-")
    return cleaned[:80] or "packet"


def load_jobs() -> Dict[str, Dict[str, Any]]:
    if not JOBS_FILE.exists():
        return {}
    try:
        return json.loads(JOBS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def load_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def save_json_file(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(path)


def normalize_email(email: str) -> str:
    return email.strip().lower()


def hash_password(password: str, salt: Optional[str] = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 180_000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        method, salt, digest = stored.split("$", 2)
    except ValueError:
        return False
    if method != "pbkdf2_sha256":
        return False
    expected = hash_password(password, salt).split("$", 2)[2]
    return hmac.compare_digest(expected, digest)


def load_users() -> Dict[str, Dict[str, Any]]:
    return load_json_file(USERS_FILE, {})


def save_users(users: Dict[str, Dict[str, Any]]) -> None:
    save_json_file(USERS_FILE, users)


def bootstrap_admin_from_env() -> None:
    users = load_users()
    if users:
        return
    email = normalize_email(os.environ.get("ADMIN_EMAIL", ""))
    password = os.environ.get("ADMIN_PASSWORD", "")
    if not email or not password:
        return
    users[email] = {
        "email": email,
        "name": os.environ.get("ADMIN_NAME", "Admin").strip() or "Admin",
        "role": "Admin",
        "active": True,
        "password_hash": hash_password(password),
        "created_at": utc_now(),
        "updated_at": utc_now(),
    }
    save_users(users)


def create_user(email: str, password: str, role: str, name: str = "", active: bool = True) -> Dict[str, Any]:
    clean_email = normalize_email(email)
    if not clean_email or "@" not in clean_email:
        raise HTTPException(status_code=400, detail="Enter a valid email address.")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters.")
    if role not in ROLE_ORDER:
        raise HTTPException(status_code=400, detail="Unsupported role.")
    users = load_users()
    if clean_email in users:
        raise HTTPException(status_code=400, detail="User already exists.")
    user = {
        "email": clean_email,
        "name": name.strip() or clean_email,
        "role": role,
        "active": active,
        "password_hash": hash_password(password),
        "created_at": utc_now(),
        "updated_at": utc_now(),
    }
    users[clean_email] = user
    save_users(users)
    return user


def load_sessions() -> Dict[str, Dict[str, Any]]:
    return load_json_file(SESSIONS_FILE, {})


def save_sessions(sessions: Dict[str, Dict[str, Any]]) -> None:
    save_json_file(SESSIONS_FILE, sessions)


def create_session(email: str) -> str:
    sessions = load_sessions()
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=AUTH_SESSION_DAYS)
    sessions[token] = {
        "email": normalize_email(email),
        "created_at": utc_now(),
        "expires_at": expires_at.isoformat(timespec="seconds"),
    }
    save_sessions(sessions)
    return token


def active_admin_count(users: Dict[str, Dict[str, Any]]) -> int:
    return sum(1 for user in users.values() if user.get("active", True) and user.get("role") == "Admin")


def expire_session(token: str) -> None:
    if not token:
        return
    sessions = load_sessions()
    if token in sessions:
        del sessions[token]
        save_sessions(sessions)


def current_user_from_request(request: Request) -> Optional[Dict[str, Any]]:
    token = request.cookies.get(SESSION_COOKIE, "")
    if not token:
        return None
    sessions = load_sessions()
    session = sessions.get(token)
    if not session:
        return None
    try:
        expires_at = datetime.fromisoformat(session.get("expires_at", ""))
    except ValueError:
        expire_session(token)
        return None
    if expires_at < datetime.now(timezone.utc):
        expire_session(token)
        return None
    user = load_users().get(normalize_email(session.get("email", "")))
    if not user or not user.get("active", True):
        return None
    return {k: v for k, v in user.items() if k != "password_hash"}


def role_at_least(user: Optional[Dict[str, Any]], role: str) -> bool:
    if not user:
        return False
    return ROLE_ORDER.get(user.get("role", "Viewer"), 0) >= ROLE_ORDER[role]


def actor_from_request(request: Request) -> str:
    user = getattr(request.state, "user", None) or {}
    return user.get("email") or "system"


def safe_next_path(next_url: str) -> str:
    next_url = (next_url or "/").strip()
    if not next_url.startswith("/") or next_url.startswith("//"):
        return "/"
    return next_url


def route_min_role(path: str, method: str) -> str:
    if path.startswith(("/settings", "/templates", "/users", "/diagnostics", "/learned-rules", "/backups")):
        return "Admin"
    if method.upper() == "POST" and path.endswith("/delete"):
        return "Admin"
    if method.upper() == "POST":
        return "Reviewer"
    return "Viewer"


def save_jobs(jobs: Dict[str, Dict[str, Any]]) -> None:
    tmp = JOBS_FILE.with_suffix(".tmp")
    tmp.write_text(json.dumps(jobs, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(JOBS_FILE)


def load_form_jobs() -> Dict[str, Dict[str, Any]]:
    return load_json_file(FORM_JOBS_FILE, {})


def save_form_jobs(jobs: Dict[str, Dict[str, Any]]) -> None:
    save_json_file(FORM_JOBS_FILE, jobs)


def get_form_job(form_job_id: str) -> Dict[str, Any]:
    job = load_form_jobs().get(form_job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Form verification not found")
    return job


def update_form_job(form_job_id: str, **changes: Any) -> Dict[str, Any]:
    jobs = load_form_jobs()
    job = jobs.get(form_job_id, {})
    job.update(changes)
    job["updated_at"] = utc_now()
    jobs[form_job_id] = job
    save_form_jobs(jobs)
    return job


def get_job(job_id: str) -> Dict[str, Any]:
    jobs = load_jobs()
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


def update_job(job_id: str, **changes: Any) -> Dict[str, Any]:
    with jobs_lock:
        jobs = load_jobs()
        job = jobs.get(job_id, {})
        job.update(changes)
        job["updated_at"] = utc_now()
        jobs[job_id] = job
        save_jobs(jobs)
    return job


def add_notification(
    title: str,
    detail: str,
    *,
    recipient: str = "",
    url: str = "",
    kind: str = "info",
) -> Dict[str, Any]:
    items = load_json_file(NOTIFICATIONS_FILE, [])
    item = {
        "id": uuid4().hex[:12],
        "title": title,
        "detail": detail,
        "recipient": normalize_email(recipient) if recipient else "",
        "url": url,
        "kind": kind,
        "read_by": [],
        "created_at": utc_now(),
    }
    items.append(item)
    save_json_file(NOTIFICATIONS_FILE, items[-1000:])
    if recipient and os.environ.get("SMTP_HOST"):
        try:
            message = EmailMessage()
            message["Subject"] = f"{APP_NAME}: {title}"
            message["From"] = os.environ.get("SMTP_FROM", os.environ.get("SMTP_USER", ""))
            message["To"] = recipient
            message.set_content(f"{detail}\n\n{url}".strip())
            port = int(os.environ.get("SMTP_PORT", "587"))
            with smtplib.SMTP(os.environ["SMTP_HOST"], port, timeout=20) as server:
                if os.environ.get("SMTP_STARTTLS", "true").lower() in {"1", "true", "yes"}:
                    server.starttls()
                if os.environ.get("SMTP_USER"):
                    server.login(os.environ["SMTP_USER"], os.environ.get("SMTP_PASSWORD", ""))
                server.send_message(message)
        except Exception:
            pass
    return item


def user_notifications(user: Dict[str, Any]) -> List[Dict[str, Any]]:
    email = normalize_email(user.get("email", ""))
    is_admin = user.get("role") == "Admin"
    return [
        item for item in reversed(load_json_file(NOTIFICATIONS_FILE, []))
        if not item.get("recipient") or item.get("recipient") == email or (is_admin and item.get("recipient") == "admin")
    ]


def production_calendar() -> Dict[str, Any]:
    return load_json_file(PRODUCTION_CALENDAR_FILE, {"non_production_dates": {}, "notes": {}})


def auto_file_form(job: Dict[str, Any]) -> Path:
    period = (job.get("period") or datetime.now().strftime("%Y-%m")).strip()
    year, month = (period.split("-", 1) + ["unknown"])[:2]
    target_dir = FORM_FILING_DIR / slugify(year) / slugify(month) / slugify(job.get("template_code", "form"))
    target_dir.mkdir(parents=True, exist_ok=True)
    source = Path(job["input_path"])
    target = target_dir / f"{job['id']}_{slugify(source.name)}"
    shutil.copy2(source, target)
    output_dir = Path(job["output_dir"])
    for filename in ("verification.json", "issues.csv"):
        artifact = output_dir / filename
        if artifact.exists():
            shutil.copy2(artifact, target_dir / f"{job['id']}_{filename}")
    update_form_job(job["id"], filed_path=str(target), filed_at=utc_now())
    return target


def create_backup_archive() -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    archive = BACKUPS_DIR / f"sqr-backup-{timestamp}.zip"
    excluded = {BACKUPS_DIR.resolve(), SESSIONS_FILE.resolve()}
    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as handle:
        for path in DATA_DIR.rglob("*"):
            resolved = path.resolve()
            if not path.is_file() or resolved == SESSIONS_FILE.resolve():
                continue
            if BACKUPS_DIR.resolve() in resolved.parents:
                continue
            handle.write(path, path.relative_to(DATA_DIR))
    webhook = os.environ.get("BACKUP_WEBHOOK_URL", "").strip()
    if webhook:
        request = urllib.request.Request(
            webhook,
            data=archive.read_bytes(),
            headers={
                "Content-Type": "application/zip",
                "X-Backup-Filename": archive.name,
                **({"Authorization": f"Bearer {os.environ['BACKUP_WEBHOOK_TOKEN']}"} if os.environ.get("BACKUP_WEBHOOK_TOKEN") else {}),
            },
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=180):
            pass
    return archive


def issue_key(job_id: str, issue: Dict[str, Any]) -> str:
    return f"{job_id}:{issue.get('key') or issue.get('id')}"


def flags_only_summary(summary: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not summary:
        return summary
    clean = dict(summary)
    clean["info_count"] = 0
    flags = [item for item in clean.get("issues", []) if item.get("check_status") == "fail"]
    active_flags = [item for item in flags if item.get("status", "Open") not in RESOLVED_ISSUE_STATUSES]
    reviewed_flags = [item for item in flags if item.get("status", "Open") in RESOLVED_ISSUE_STATUSES]
    clean["issues"] = active_flags
    clean["reviewed_issues"] = reviewed_flags
    clean["active_fail_count"] = len(active_flags)
    clean["reviewed_fail_count"] = len(reviewed_flags)
    clean["fail_count"] = len(active_flags)
    clean["pass_count"] = int(clean.get("pass_count") or 0) + len(reviewed_flags)
    if flags and not active_flags:
        clean["overall"] = "PASS"
    clean["failures"] = [item for item in clean.get("failures", [])]
    return clean


def apply_board_review(job: Dict[str, Any]) -> Dict[str, Any]:
    board = load_json_file(BOARD_FILE, {})
    summary = job.get("summary") or {}
    for issue in summary.get("issues", []) + summary.get("reviewed_issues", []):
        state = board.get(issue_key(job["id"], issue), {})
        if state:
            issue.update({k: v for k, v in state.items() if k in {
                "status", "comment", "assignee", "due_date", "priority", "updated_at"
            }})
    job["summary"] = flags_only_summary(summary)
    return job


def public_job(job: Dict[str, Any]) -> Dict[str, Any]:
    clean = dict(job)
    clean = apply_board_review(clean)
    return clean


def audit_event(job: Dict[str, Any], action: str, detail: str, **extra: Any) -> Dict[str, Any]:
    events = list(job.get("audit_trail") or [])
    event = {
        "id": f"AUD-{len(events) + 1:04d}",
        "at": utc_now(),
        "user": extra.pop("user", "system"),
        "action": action,
        "detail": detail,
    }
    event.update(extra)
    events.append(event)
    job["audit_trail"] = events
    return job


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    bootstrap_admin_from_env()
    path = request.url.path
    public_path = (
        path == "/healthz"
        or path.startswith("/static/")
        or path in {"/login", "/logout", "/setup"}
    )
    user = current_user_from_request(request)
    request.state.user = user

    if not public_path and not load_users():
        return RedirectResponse(url="/setup", status_code=303)
    if not public_path and not user:
        next_path = safe_next_path(path)
        if request.url.query:
            next_path = f"{next_path}?{request.url.query}"
        return RedirectResponse(url=f"/login?next={next_path}", status_code=303)
    if not public_path and not role_at_least(user, route_min_role(path, request.method)):
        return HTMLResponse("You do not have permission to access this page.", status_code=403)

    response = await call_next(request)
    return response


def config_file_path(config_name: str) -> Path:
    filename = CONFIG_EDITABLE_FILES.get(config_name)
    if not filename:
        raise HTTPException(status_code=404, detail="Config file not found")
    path = (CONFIG_DIR / filename).resolve()
    if CONFIG_DIR not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="Config file not found")
    return path


def load_config_panels(selected: str = "rules", message: str = "", error: str = "") -> Dict[str, Any]:
    panels = []
    for key, filename in CONFIG_EDITABLE_FILES.items():
        path = config_file_path(key)
        panels.append(
            {
                "key": key,
                "filename": filename,
                "selected": key == selected,
                "content": path.read_text(encoding="utf-8"),
            }
        )
    return {
        "panels": panels,
        "selected": selected if selected in CONFIG_EDITABLE_FILES else "rules",
        "message": message,
        "error": error,
    }


def rules_config() -> Dict[str, Any]:
    return yaml.safe_load(config_file_path("rules").read_text(encoding="utf-8")) or {}


def save_rules_config(data: Dict[str, Any]) -> None:
    path = config_file_path("rules")
    backup = path.with_suffix(path.suffix + f".{utc_now().replace(':', '-')}.bak")
    shutil.copy2(path, backup)
    path.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")


def get_form_templates() -> Dict[str, Dict[str, Any]]:
    custom = load_json_file(TEMPLATES_FILE, {})
    data = rules_config()
    expected = (
        data.get("rules", {})
        .get("field_coverage_audit", {})
        .get("expected_fields_by_form", {})
        or {}
    )
    templates = {}
    for code, fields in expected.items():
        templates[code] = {
            "code": code,
            "label": custom.get(code, {}).get("label", _human_form_label(code)),
            "fields": list(fields or []),
            "active": custom.get(code, {}).get("active", True),
            "category": custom.get(code, {}).get("category", "Packet"),
            "required_terms": custom.get(code, {}).get("required_terms", []),
            "require_date": custom.get(code, {}).get("require_date", False),
            "required_signatures": int(custom.get(code, {}).get("required_signatures", 0) or 0),
            "regions": custom.get(code, {}).get("regions", []),
            "sample_filename": custom.get(code, {}).get("sample_filename", ""),
            "frequency": custom.get(code, {}).get("frequency", "As needed"),
            "tracking_enabled": custom.get(code, {}).get("tracking_enabled", False),
            "related_templates": custom.get(code, {}).get("related_templates", []),
        }
    for code, item in custom.items():
        templates.setdefault(
            code,
            {
                "code": code,
                "label": item.get("label", _human_form_label(code)),
                "fields": item.get("fields", []),
                "active": item.get("active", True),
                "category": item.get("category", "Packet"),
                "required_terms": item.get("required_terms", []),
                "require_date": item.get("require_date", False),
                "required_signatures": int(item.get("required_signatures", 0) or 0),
                "regions": item.get("regions", []),
                "sample_filename": item.get("sample_filename", ""),
                "frequency": item.get("frequency", "As needed"),
                "tracking_enabled": item.get("tracking_enabled", False),
                "related_templates": item.get("related_templates", []),
            },
        )
    for code, defaults in DEFAULT_FOOD_SAFETY_TEMPLATES.items():
        templates.setdefault(code, {
            "code": code,
            "label": defaults["label"],
            "fields": [],
            "active": True,
            "category": "Food Safety",
            "required_terms": defaults["required_terms"],
            "require_date": defaults["require_date"],
            "required_signatures": defaults["required_signatures"],
            "regions": [],
            "sample_filename": "",
            "frequency": "Daily" if code == "FORKLIFT_DAILY" else "Monthly",
            "tracking_enabled": True,
            "related_templates": [],
        })
    return dict(sorted(templates.items()))


def _human_form_label(code: str) -> str:
    return code.replace("_", " ").title()


def save_form_template(
    code: str,
    label: str,
    fields_text: str,
    active: bool,
    category: str = "Packet",
    required_terms_text: str = "",
    require_date: bool = False,
    required_signatures: int = 0,
    regions_text: str = "",
    sample_filename: str = "",
    frequency: str = "As needed",
    tracking_enabled: bool = False,
    related_templates_text: str = "",
) -> str:
    clean_code = re.sub(r"[^A-Za-z0-9_]+", "_", code.strip().upper()).strip("_")
    if not clean_code:
        raise HTTPException(status_code=400, detail="Form code is required.")
    fields = [re.sub(r"[^A-Za-z0-9_]+", "_", item.strip().lower()).strip("_")
              for item in re.split(r"[,\n]+", fields_text)]
    fields = [item for item in fields if item]
    required_terms = [item.strip() for item in re.split(r"[,\n]+", required_terms_text) if item.strip()]
    try:
        regions = json.loads(regions_text or "[]")
        if not isinstance(regions, list):
            raise ValueError
    except (json.JSONDecodeError, ValueError):
        raise HTTPException(status_code=400, detail="Visual regions must be a JSON list.")
    templates = get_form_templates()
    previous = templates.get(clean_code, {})
    templates[clean_code] = {
        "code": clean_code,
        "label": label.strip() or _human_form_label(clean_code),
        "fields": fields,
        "active": active,
        "category": category if category in {"Packet", "Food Safety"} else "Packet",
        "required_terms": required_terms,
        "require_date": require_date,
        "required_signatures": max(0, int(required_signatures or 0)),
        "regions": regions,
        "sample_filename": sample_filename or previous.get("sample_filename", ""),
        "frequency": frequency if frequency in {"Daily", "Weekly", "Monthly", "As needed"} else "As needed",
        "tracking_enabled": tracking_enabled,
        "related_templates": [
            re.sub(r"[^A-Za-z0-9_]+", "_", item.strip().upper()).strip("_")
            for item in re.split(r"[,\n]+", related_templates_text)
            if item.strip()
        ],
    }
    save_json_file(TEMPLATES_FILE, templates)
    data = rules_config()
    coverage = data.setdefault("rules", {}).setdefault("field_coverage_audit", {})
    expected = coverage.setdefault("expected_fields_by_form", {})
    expected[clean_code] = fields
    save_rules_config(data)
    return clean_code


def missing_form_expectations(month: str) -> List[Dict[str, Any]]:
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        start = datetime.now().date().replace(day=1)
        month = start.strftime("%Y-%m")
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    jobs = list(load_form_jobs().values())
    calendar = production_calendar()
    closed = set(calendar.get("non_production_dates", {}))
    missing: List[Dict[str, Any]] = []
    for template in get_form_templates().values():
        if template.get("category") != "Food Safety" or not template.get("active", True) or not template.get("tracking_enabled"):
            continue
        frequency = template.get("frequency", "As needed")
        expected: List[str] = []
        cursor = start
        if frequency == "Daily":
            while cursor < next_month:
                date_key = cursor.isoformat()
                if cursor.weekday() < 5 and date_key not in closed:
                    expected.append(date_key)
                cursor += timedelta(days=1)
        elif frequency == "Weekly":
            while cursor < next_month:
                if cursor.weekday() == 0 and cursor.isoformat() not in closed:
                    expected.append(cursor.isoformat())
                cursor += timedelta(days=1)
        elif frequency == "Monthly":
            expected.append(month)
        for due in expected:
            found = any(
                job.get("template_code") == template["code"]
                and (
                    job.get("form_date") == due
                    or job.get("period") == due
                    or (frequency in {"Daily", "Weekly"} and str(job.get("period", "")).startswith(month)
                        and due in str(job.get("name", "")))
                )
                for job in jobs
            )
            if not found:
                missing.append({
                    "template_code": template["code"],
                    "template_label": template["label"],
                    "frequency": frequency,
                    "due": due,
                })
    return missing


def issue_board_items() -> List[Dict[str, Any]]:
    board = load_json_file(BOARD_FILE, {})
    items: List[Dict[str, Any]] = []
    for job in sorted([public_job(j) for j in load_jobs().values()], key=lambda j: j.get("created_at", ""), reverse=True):
        summary = job.get("summary") or {}
        for issue in summary.get("issues", []) + summary.get("reviewed_issues", []):
            key = issue_key(job["id"], issue)
            state = board.get(key, {})
            items.append({
                "board_key": key,
                "job_id": job["id"],
                "packet_name": job.get("packet_name"),
                "created_at": job.get("created_at"),
                **issue,
                **state,
                "history": state.get("history", []),
            })
    return items


def issue_counts(items: List[Dict[str, Any]]) -> Dict[str, int]:
    counts = {"open": 0, "corrected": 0, "resolved": 0, "total": len(items)}
    for item in items:
        status = (item.get("status") or "Open").lower()
        if status == "corrected":
            counts["corrected"] += 1
        elif status in {"resolved", "accepted as is", "false positive"}:
            counts["resolved"] += 1
        else:
            counts["open"] += 1
    return counts


def issue_type(name: str, detail: str) -> str:
    text = f"{name} {detail}".lower()
    if "required form" in text or "present" in text and "not detected" in text:
        return "Missing Document"
    if "field coverage" in text or "not detected" in text:
        return "Missing Field"
    if "cross-page" in text or "≠" in text or "disagrees" in text:
        return "Mismatch"
    if "weight calc" in text or "case count" in text or "total" in text:
        return "Math Error"
    if "signature" in text or "initial" in text:
        return "Missing Initials/Signature"
    if "spec" in text or "outside" in text:
        return "Out of Spec"
    if "defect" in text or "photo" in text or "evidence" in text:
        return "Missing Evidence"
    if "metal" in text:
        return "Metal Detector"
    return "Review"


def approved_false_positive_rule(check: Any, report: Any) -> Optional[Dict[str, Any]]:
    name = re.sub(r"\(p\d+\)", "(p#)", check.name.lower())
    detail = check.detail.lower()
    for rule in load_json_file(LEARNED_RULES_FILE, []):
        if rule.get("status") != "Approved":
            continue
        if rule.get("name_pattern") and rule["name_pattern"] not in name:
            continue
        if rule.get("detail_pattern") and rule["detail_pattern"] not in detail:
            continue
        form_code = rule.get("form_code")
        if form_code:
            page_codes = {
                page.form_code for page in report.pages
                if page.page_no in (check.pages or [])
            }
            if form_code not in page_codes:
                continue
        return rule
    return None


def render_pdf_page_to_png(pdf_path: Path, page_index: int, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        return output_path
    try:
        import fitz

        with fitz.open(str(pdf_path)) as doc:
            if page_index < 0 or page_index >= doc.page_count:
                raise HTTPException(status_code=404, detail="PDF page not found")
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=fitz.Matrix(1.7, 1.7), alpha=False)
            pix.save(str(output_path))
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Could not render comparison page: {exc}") from exc
    return output_path


def build_issues(report: Any, existing_reviews: Optional[Dict[str, Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    reviews = existing_reviews or {}
    issues: List[Dict[str, Any]] = []
    for idx, check in enumerate([c for c in report.all_checks if c.status == "fail"], start=1):
        stable_key = "|".join(
            [
                check.status,
                str(None if check.sub_packet is None else check.sub_packet + 1),
                check.name,
                check.detail,
                ",".join(map(str, check.pages or [])),
            ]
        )
        review = reviews.get(stable_key, {})
        learned_rule = approved_false_positive_rule(check, report)
        if learned_rule and not review:
            review = {
                "status": "False Positive",
                "comment": f"Automatically resolved by approved rule {learned_rule.get('id')}.",
                "updated_at": utc_now(),
            }
        severity = "High"
        issues.append(
            {
                "id": f"ISS-{idx:03d}",
                "key": stable_key,
                "severity": severity,
                "status": review.get("status", "Open"),
                "issue_type": issue_type(check.name, check.detail),
                "check_status": check.status,
                "name": check.name,
                "detail": check.detail,
                "pages": check.pages,
                "sub_packet": None if check.sub_packet is None else check.sub_packet + 1,
                "comment": review.get("comment", ""),
                "updated_at": review.get("updated_at"),
                "learned_rule_id": learned_rule.get("id") if learned_rule else "",
            }
        )
    return issues


def merge_issue_reviews(summary: Dict[str, Any], reviews: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    review_map = reviews or {}
    for issue in summary.get("issues", []):
        if issue["key"] in review_map:
            issue.update({k: v for k, v in review_map[issue["key"]].items() if k in {"status", "comment", "updated_at"}})
    return summary


def summarize_report(report: Any) -> Dict[str, Any]:
    issues = build_issues(report)
    return {
        "overall": report.overall,
        "pass_count": report.n_pass,
        "fail_count": report.n_fail,
        "info_count": 0,
        "page_count": len(report.pages),
        "sub_packet_count": len(report.sub_packets),
        "sub_packets": [
            {
                "index": sp.index + 1,
                "wo": sp.primary_wo,
                "po": sp.primary_po,
                "customer": sp.primary_customer,
                "product": sp.primary_product,
                "cases": sp.cases,
                "total_lbs": sp.total_lbs,
            }
            for sp in report.sub_packets
        ],
        "failures": [
            {
                "name": c.name,
                "detail": c.detail,
                "pages": c.pages,
                "sub_packet": None if c.sub_packet is None else c.sub_packet + 1,
            }
            for c in report.all_checks
            if c.status == "fail"
        ],
        "issues": issues,
        "warnings": [],
    }


def output_files(job: Dict[str, Any]) -> List[Dict[str, Any]]:
    out_dir = Path(job["output_dir"])
    packet_name = job["packet_name"]
    candidates = [
        ("Verified PDF", out_dir / f"{packet_name}_AI_VERIFIED.pdf"),
        ("Issues CSV", out_dir / f"{packet_name}_issues.csv"),
        ("Trace JSON", out_dir / f"{packet_name}_trace.json"),
        ("Cross-reference Matrix", out_dir / f"{packet_name}_cross_reference_matrix.xlsx"),
        ("Summary PNG", out_dir / f"{packet_name}_summary.png"),
    ]
    return [
        {
            "label": label,
            "filename": path.name,
            "url": f"/jobs/{job['id']}/download/{path.name}",
            "download_url": f"/jobs/{job['id']}/download/{path.name}",
            "preview_url": f"/jobs/{job['id']}/view/{path.name}",
            "table_preview_url": f"/jobs/{job['id']}/preview/{path.name}",
            "extension": path.suffix.lower().lstrip("."),
            "previewable": path.suffix.lower() in {".pdf", ".png", ".csv", ".json", ".xlsx"},
            "table_preview": path.suffix.lower() in {".csv", ".xlsx"},
        }
        for label, path in candidates
        if path.exists()
    ]


def output_file_path(job: Dict[str, Any], filename: str) -> Path:
    out_dir = Path(job["output_dir"]).resolve()
    path = (out_dir / filename).resolve()
    if out_dir not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return path


def page_catalog(job: Dict[str, Any]) -> List[Dict[str, Any]]:
    trace_path = Path(job["output_dir"]) / f"{job['packet_name']}_trace.json"
    if not trace_path.exists():
        candidates = list(Path(job["output_dir"]).glob("*_trace.json"))
        trace_path = candidates[0] if candidates else trace_path
    trace = load_json_file(trace_path, {})
    pages = []
    label_counts: Dict[str, int] = {}
    for page in trace.get("pages", []) or []:
        page_no = int(page.get("page_no") or 0)
        if page_no < 1:
            continue
        label = (
            page.get("form_label")
            or page.get("form_type")
            or page.get("form_code")
            or "Unidentified form"
        )
        label_counts[label] = label_counts.get(label, 0) + 1
        pages.append({
            "page_no": page_no,
            "label": label,
            "form_code": page.get("form_code") or "",
            "display": f"Page {page_no} - {label}",
            "occurrence": label_counts[label],
        })
    if not pages:
        count = int((job.get("summary") or {}).get("page_count") or 0)
        pages = [
            {"page_no": page_no, "label": "Packet page", "form_code": "", "display": f"Page {page_no}"}
            for page_no in range(1, count + 1)
        ]
    return pages


def packet_workflow_counts(jobs: List[Dict[str, Any]]) -> Dict[str, int]:
    counts = {"total": len(jobs), "pending": 0, "correction": 0, "approved": 0}
    for job in jobs:
        status = (job.get("workflow_status") or "AI Processing").lower()
        if status == "approved":
            counts["approved"] += 1
        elif status == "correction required":
            counts["correction"] += 1
        else:
            counts["pending"] += 1
    return counts


def derive_packet_name(report: Any, fallback: str) -> str:
    invoice = next(
        (str(page.fields.get("invoice_no")) for page in report.pages if page.fields.get("invoice_no")),
        "",
    )
    sub_packet = report.sub_packets[0] if report.sub_packets else None
    customer = getattr(sub_packet, "primary_customer", "") if sub_packet else ""
    wo = getattr(sub_packet, "primary_wo", "") if sub_packet else ""
    parts = [item for item in [invoice, customer, f"WO-{wo}" if wo else ""] if item]
    return slugify("-".join(parts) if parts else fallback)


def rename_job_outputs(job: Dict[str, Any], new_name: str) -> None:
    old_name = job["packet_name"]
    if not new_name or new_name == old_name:
        return
    out_dir = Path(job["output_dir"])
    for path in list(out_dir.glob(f"{old_name}*")):
        path.rename(path.with_name(new_name + path.name[len(old_name):]))
    old_work = out_dir / "_work" / old_name
    new_work = out_dir / "_work" / new_name
    if old_work.exists() and not new_work.exists():
        old_work.rename(new_work)


def page_image_path(job: Dict[str, Any], page_no: int) -> Path:
    if page_no < 1:
        raise HTTPException(status_code=404, detail="Page not found")
    out_dir = Path(job["output_dir"]).resolve()
    packet_name = job["packet_name"]
    candidates = [
        out_dir / "annotated_pages" / f"p-{page_no:02d}_annot.png",
        out_dir / "_work" / packet_name / "pages" / f"p-{page_no:02d}.png",
    ]
    for candidate in candidates:
        path = candidate.resolve()
        if out_dir in path.parents and path.exists():
            return path
    raise HTTPException(status_code=404, detail="Page image not found")


def _cell_to_preview(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def preview_table_file(path: Path) -> Dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows: List[List[Any]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= PREVIEW_MAX_ROWS:
                    break
                rows.append([_cell_to_preview(cell) for cell in row[:PREVIEW_MAX_COLS]])
        return {
            "filename": path.name,
            "type": "csv",
            "truncated": len(rows) >= PREVIEW_MAX_ROWS,
            "sheets": [{"name": "CSV", "rows": rows}],
        }
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= PREVIEW_MAX_ROWS:
                    break
                rows.append([_cell_to_preview(cell) for cell in row[:PREVIEW_MAX_COLS]])
            sheets.append(
                {
                    "name": ws.title,
                    "rows": rows,
                    "source_rows": ws.max_row,
                    "source_cols": ws.max_column,
                }
            )
        return {
            "filename": path.name,
            "type": "xlsx",
            "truncated": any(
                sheet.get("source_rows", 0) > PREVIEW_MAX_ROWS
                or sheet.get("source_cols", 0) > PREVIEW_MAX_COLS
                for sheet in sheets
            ),
            "sheets": sheets,
        }
    raise HTTPException(status_code=415, detail="Table preview is only available for CSV and XLSX files")


def write_reviewed_issues_csv(job: Dict[str, Any]) -> None:
    summary = (job.get("summary") or {})
    path = Path(job["output_dir"]) / f"{job['packet_name']}_issues.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["#", "Status", "Sub-packet", "Check", "Detail", "Pages", "Resolution (human)"])
        for idx, issue in enumerate(summary.get("issues", []), start=1):
            writer.writerow([
                idx,
                "FLAG",
                issue.get("sub_packet", ""),
                issue.get("name", ""),
                issue.get("detail", ""),
                ", ".join(map(str, issue.get("pages") or [])),
                issue.get("comment", ""),
            ])


def write_reviewed_summary_png(job: Dict[str, Any]) -> Optional[Path]:
    summary = job.get("summary") or {}
    out_path = Path(job["output_dir"]) / f"{job['packet_name']}_summary.png"
    try:
        from PIL import Image, ImageDraw, ImageFont

        im = Image.new("RGB", (2100, 2700), "white")
        draw = ImageDraw.Draw(im)
        try:
            big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 64)
            h1 = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 40)
            bd = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
            rg = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 24)
        except Exception:
            big = h1 = bd = rg = ImageFont.load_default()

        overall = summary.get("overall", "PASS")
        banner = (200, 240, 200) if overall == "PASS" else (255, 210, 200)
        draw.rectangle([(0, 0), (2100, 200)], fill=banner)
        draw.text((40, 30), f"AI VERIFICATION - {overall}", fill=(0, 0, 0), font=big)
        draw.text((40, 110), f"Packet: {job['packet_name']}", fill=(0, 0, 0), font=h1)

        y = 250
        rows = [
            ("Sub-packets:", str(summary.get("sub_packet_count", 0))),
            ("Customer:", (summary.get("sub_packets") or [{}])[0].get("customer", "(unknown)") if summary.get("sub_packets") else "(unknown)"),
            ("Total pages:", str(summary.get("page_count", ""))),
            ("Verified at:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for key, value in rows:
            draw.text((60, y), key, fill=(0, 0, 0), font=bd)
            draw.text((360, y), str(value), fill=(0, 0, 0), font=rg)
            y += 50

        y += 30
        draw.text((60, y), "Result tally:", fill=(0, 0, 0), font=h1)
        y += 60
        draw.text((80, y), f"{summary.get('pass_count', 0)} checks passed", fill=(40, 140, 60), font=bd)
        y += 50
        draw.text((80, y), f"{summary.get('fail_count', 0)} flag(s) raised", fill=(235, 130, 25), font=bd)
        y += 70

        issues = summary.get("issues") or []
        if issues:
            draw.text((60, y), "Issues to review:", fill=(235, 130, 25), font=h1)
            y += 60
            for idx, issue in enumerate(issues, start=1):
                sp_tag = f" [sub-packet {issue.get('sub_packet')}]" if issue.get("sub_packet") else ""
                line = f"{idx}. {issue.get('name', '')}{sp_tag} - {issue.get('detail', '')}"
                if len(line) > 160:
                    line = line[:157] + "..."
                draw.text((80, y), line, fill=(0, 0, 0), font=rg)
                y += 36
                if y > 2460:
                    break
        else:
            draw.text((60, y), "No issues raised. Packet is accepted after review.", fill=(40, 140, 60), font=h1)

        sig_y = 2420
        draw.text((60, sig_y), "Reviewer signature: _____________________________   Date: ___________", fill=(0, 0, 0), font=bd)
        draw.text((60, sig_y + 60), "Accepted/false-positive/resolved flags are counted as passed by reviewer.", fill=(80, 80, 80), font=rg)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        im.save(out_path)
        return out_path
    except Exception as exc:  # noqa: BLE001
        print(f"Reviewed summary image skipped: {exc}")
        return None


def replace_pdf_summary_page(job: Dict[str, Any], summary_png: Path) -> None:
    pdf_path = Path(job["output_dir"]) / f"{job['packet_name']}_AI_VERIFIED.pdf"
    if not pdf_path.exists() or not summary_png.exists():
        return
    try:
        from PIL import Image
        from pypdf import PdfReader, PdfWriter

        summary_pdf = summary_png.with_suffix(".reviewed-summary.pdf")
        Image.open(summary_png).convert("RGB").save(summary_pdf)
        summary_reader = PdfReader(str(summary_pdf))
        original_reader = PdfReader(str(pdf_path))
        writer = PdfWriter()
        writer.add_page(summary_reader.pages[0])
        for page in list(original_reader.pages)[1:]:
            writer.add_page(page)
        tmp = pdf_path.with_suffix(".reviewed.tmp.pdf")
        with tmp.open("wb") as f:
            writer.write(f)
        tmp.replace(pdf_path)
    except Exception as exc:  # noqa: BLE001
        print(f"Reviewed PDF summary replacement skipped: {exc}")


def refresh_reviewed_output_files(raw_job: Dict[str, Any]) -> None:
    if raw_job.get("status") != "complete" or not raw_job.get("summary"):
        return
    job = public_job(raw_job)
    write_reviewed_issues_csv(job)
    summary_png = write_reviewed_summary_png(job)
    if summary_png:
        replace_pdf_summary_page(job, summary_png)


def read_trace(job: Dict[str, Any]) -> Dict[str, Any]:
    trace_path = Path(job["output_dir"]) / f"{job['packet_name']}_trace.json"
    if not trace_path.exists():
        return {}
    return load_json_file(trace_path, {})


def field_confidence_rows(job: Dict[str, Any]) -> List[Dict[str, Any]]:
    trace = read_trace(job)
    rows: List[Dict[str, Any]] = []
    for page in trace.get("pages", []) or []:
        fields = page.get("fields", {}) or {}
        confidence = float(page.get("confidence_estimate") or fields.get("_page_confidence") or 0)
        field_conf = fields.get("_field_confidence") or {}
        for key, value in fields.items():
            if key.startswith("_") or value in (None, "", [], {}):
                continue
            if isinstance(value, (list, dict)):
                continue
            rows.append({
                "page_no": page.get("page_no"),
                "form": page.get("form_label") or page.get("form_code"),
                "field": key,
                "value": value,
                "confidence": round(float(field_conf.get(key, confidence) or confidence), 3),
                "backend": page.get("ocr_backend_used"),
            })
    return rows


def write_audit_workbook(path: Path, month: str, jobs: List[Dict[str, Any]], inspections: List[Dict[str, Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Packet Summary"
    headers = ["Packet", "Uploaded", "Status", "Overall", "Flags", "Open flags", "Signed by", "Decision"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFEAEAEA")
    for job in jobs:
        summary = job.get("summary") or {}
        issues = summary.get("issues") or []
        open_flags = sum(1 for item in issues if (item.get("status") or "Open") == "Open")
        signoff = job.get("review_signoff") or {}
        ws.append([
            job.get("packet_name"),
            job.get("created_at"),
            job.get("status"),
            summary.get("overall", ""),
            summary.get("fail_count", 0),
            open_flags,
            signoff.get("reviewer_name", ""),
            signoff.get("decision", ""),
        ])
    ws2 = wb.create_sheet("Live Inspections")
    ws2.append(["Date", "Type", "Inspector", "Area", "Result", "Corrective action"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFEAEAEA")
    for item in inspections:
        ws2.append([
            item.get("date"),
            item.get("inspection_type"),
            item.get("inspector"),
            item.get("area"),
            item.get("result"),
            item.get("corrective_action"),
        ])
    ws3 = wb.create_sheet("SQF Open Items")
    ws3.append(["Packet", "Flag", "Assignee", "Due date", "Status", "Detail"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFEAEAEA")
    for item in issue_board_items():
        if (item.get("status") or "Open") in {"Resolved", "Accepted as Is", "False Positive"}:
            continue
        ws3.append([
            item.get("packet_name"),
            item.get("name"),
            item.get("assignee", ""),
            item.get("due_date", ""),
            item.get("status", "Open"),
            item.get("detail"),
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def extract_form_pdf_text(pdf_path: Path, output_dir: Path) -> List[str]:
    import fitz

    texts: List[str] = []
    output_dir.mkdir(parents=True, exist_ok=True)
    with fitz.open(str(pdf_path)) as doc:
        for index, page in enumerate(doc):
            text = page.get_text("text").strip()
            if len(text) < 80 and shutil.which("tesseract"):
                image_path = output_dir / f"page-{index + 1:03d}.png"
                page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False).save(str(image_path))
                result = subprocess.run(
                    [shutil.which("tesseract") or "tesseract", str(image_path), "stdout", "-l", "eng"],
                    capture_output=True,
                    text=True,
                    timeout=90,
                    check=False,
                )
                if result.stdout.strip():
                    text = result.stdout.strip()
            texts.append(text)
    return texts


def extract_region_text(pdf_path: Path, page_no: int, region: Dict[str, Any], output_dir: Path) -> str:
    if not shutil.which("tesseract"):
        return ""
    import fitz

    with fitz.open(str(pdf_path)) as doc:
        index = max(0, min(page_no - 1, doc.page_count - 1))
        page = doc.load_page(index)
        rect = page.rect
        x = float(region.get("x", 0)) * rect.width
        y = float(region.get("y", 0)) * rect.height
        w = float(region.get("width", 1)) * rect.width
        h = float(region.get("height", 1)) * rect.height
        clip = fitz.Rect(x, y, min(rect.width, x + w), min(rect.height, y + h))
        path = output_dir / f"region-{slugify(str(region.get('label') or 'field'))}.png"
        page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5), clip=clip, alpha=False).save(str(path))
    result = subprocess.run(
        [shutil.which("tesseract") or "tesseract", str(path), "stdout", "-l", "eng"],
        capture_output=True,
        text=True,
        timeout=60,
        check=False,
    )
    return result.stdout.strip()


def interpret_region(region: Dict[str, Any], text: str) -> tuple[str, str]:
    kind = str(region.get("interpretation") or "nonblank").lower()
    expected = str(region.get("expected") or "").strip()
    clean = text.strip()
    if kind == "date":
        ok = bool(re.search(r"\b(?:\d{1,2}[/.-]){2}(?:\d{2}|\d{4})\b", clean))
        return ("pass" if ok else "fail", f"Date detected: {clean[:120]}" if ok else "No date detected in the highlighted region")
    if kind in {"signature", "initials"}:
        ink_like = len(re.sub(r"[^A-Za-z]", "", clean)) >= 2
        return ("pass" if ink_like else "fail", f"Signature/initials detected: {clean[:120]}" if ink_like else "Signature region appears blank or unreadable")
    if kind == "checkbox":
        ok = bool(re.search(r"(?:\b(?:yes|no|ok|pass|fail|checked)\b|[xX✓✔])", clean, re.I))
        return ("pass" if ok else "fail", f"Marked choice detected: {clean[:120]}" if ok else "No marked checkbox/choice detected")
    if kind == "text" and expected:
        ok = expected.lower() in clean.lower()
        return ("pass" if ok else "fail", f"Expected text found: {expected}" if ok else f"Expected text not found: {expected}")
    return ("pass" if clean else "fail", f"Detected: {clean[:180]}" if clean else "The highlighted page region appears blank or unreadable")


def run_form_verification(form_job_id: str) -> None:
    job = update_form_job(form_job_id, status="running", message="Reading form pages")
    try:
        template = get_form_templates().get(job["template_code"])
        if not template:
            raise RuntimeError("The selected form template no longer exists.")
        pdf_path = Path(job["input_path"])
        output_dir = Path(job["output_dir"])
        pages = extract_form_pdf_text(pdf_path, output_dir / "ocr")
        combined = "\n".join(pages)
        checks: List[Dict[str, Any]] = []
        for term in template.get("required_terms", []):
            found_pages = [index + 1 for index, text in enumerate(pages) if term.lower() in text.lower()]
            checks.append({
                "name": f"Required text: {term}",
                "status": "pass" if found_pages else "fail",
                "detail": f"Found on pages {found_pages}" if found_pages else "Not detected in the uploaded form",
                "pages": found_pages,
            })
        if template.get("require_date"):
            date_matches = re.findall(r"\b(?:\d{1,2}[/.-]){2}(?:\d{2}|\d{4})\b", combined)
            checks.append({
                "name": "Date present",
                "status": "pass" if date_matches else "fail",
                "detail": f"Detected {len(date_matches)} date value(s)" if date_matches else "No date was detected",
                "pages": [],
            })
        required_signatures = int(template.get("required_signatures", 0) or 0)
        if required_signatures:
            signature_hits = len(re.findall(r"\b(?:signature|signed|verified by|initials?)\b", combined, flags=re.I))
            checks.append({
                "name": "Signature/verification coverage",
                "status": "pass" if signature_hits >= required_signatures else "fail",
                "detail": f"Detected {signature_hits}; template requires {required_signatures}",
                "pages": [],
            })
        for region in template.get("regions", []):
            if not region.get("required", True):
                continue
            region_text = extract_region_text(
                pdf_path,
                int(region.get("page", 1) or 1),
                region,
                output_dir / "regions",
            )
            status, detail = interpret_region(region, region_text)
            checks.append({
                "name": f"Required region: {region.get('label') or 'Unnamed region'}",
                "status": status,
                "detail": detail,
                "pages": [int(region.get("page", 1) or 1)],
            })
        related = [code for code in template.get("related_templates", []) if code]
        if related:
            peers = [
                item for item in load_form_jobs().values()
                if item.get("id") != form_job_id
                and item.get("period") == job.get("period")
                and item.get("template_code") in related
                and item.get("status") == "complete"
            ]
            for related_code in related:
                related_template = get_form_templates().get(related_code, {})
                matches = [item for item in peers if item.get("template_code") == related_code]
                checks.append({
                    "name": f"Related form: {related_template.get('label', related_code)}",
                    "status": "pass" if matches else "fail",
                    "detail": (
                        f"Matched {len(matches)} completed form(s) for period {job.get('period')}"
                        if matches else f"Missing related form for period {job.get('period') or 'unspecified'}"
                    ),
                    "pages": [],
                })
        if not checks:
            checks.append({
                "name": "Readable form",
                "status": "pass" if combined.strip() else "fail",
                "detail": f"Read {len(combined)} characters across {len(pages)} page(s)",
                "pages": list(range(1, len(pages) + 1)),
            })
        failures = [item for item in checks if item["status"] == "fail"]
        result = {
            "overall": "FAIL" if failures else "PASS",
            "pass_count": sum(1 for item in checks if item["status"] == "pass"),
            "fail_count": len(failures),
            "page_count": len(pages),
            "checks": checks,
        }
        save_json_file(output_dir / "verification.json", result)
        with (output_dir / "issues.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Status", "Check", "Detail", "Pages"])
            for item in failures:
                writer.writerow(["FLAG", item["name"], item["detail"], ",".join(map(str, item["pages"]))])
        update_form_job(
            form_job_id,
            status="complete",
            workflow_status="Manual Review Pending",
            completed_at=utc_now(),
            result=result,
            message="Form verification complete",
        )
        add_notification(
            "Food Safety form verification complete",
            f"{job.get('name')} finished with {len(failures)} flag(s).",
            recipient=job.get("created_by", ""),
            url=f"/forms/{form_job_id}",
            kind="warning" if failures else "success",
        )
    except Exception as exc:  # noqa: BLE001
        update_form_job(
            form_job_id,
            status="failed",
            completed_at=utc_now(),
            message=str(exc),
            error_trace=traceback.format_exc(),
        )


def run_verification(job_id: str) -> None:
    if not packet_run_semaphore.acquire(blocking=False):
        update_job(
            job_id,
            status="queued",
            message=f"Waiting for another packet to finish. Limit is {MAX_CONCURRENT_PACKET_RUNS} active packet run(s).",
            progress_percent=2,
            progress_stage="Waiting in queue",
        )
        packet_run_semaphore.acquire()
    job = update_job(
        job_id,
        status="running",
        started_at=utc_now(),
        message="Preparing packet verification",
        progress_percent=5,
        progress_stage="Preparing packet",
        processed_pages=0,
        total_pages=0,
    )
    try:
        provider = job["vision_provider"]
        if provider == "anthropic" and not os.environ.get("ANTHROPIC_API_KEY"):
            raise RuntimeError("ANTHROPIC_API_KEY is not set in the Render environment.")
        if provider == "openai" and not os.environ.get("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is not set in the Render environment.")
        cache_path: Optional[str] = str(VISION_CACHE) if provider == "mock" and VISION_CACHE.exists() else None
        last_progress = {"percent": -1, "message": ""}

        def save_progress(percent: int, message: str, processed: int, total: int) -> None:
            clean_percent = max(0, min(99, int(percent)))
            if (
                clean_percent == last_progress["percent"]
                and message == last_progress["message"]
            ):
                return
            last_progress.update(percent=clean_percent, message=message)
            update_job(
                job_id,
                progress_percent=clean_percent,
                progress_stage=message,
                message=message,
                processed_pages=processed,
                total_pages=total,
            )

        report = verify_pdf(
            pdf_path=job["input_path"],
            out_dir=job["output_dir"],
            config_dir=str(CONFIG_DIR),
            ocr_provider=provider,
            vision_cache_path=cache_path,
            packet_name=job["packet_name"],
            progress_callback=save_progress,
        )
        if not job.get("packet_name_user_supplied"):
            automatic_name = derive_packet_name(report, job["packet_name"])
            rename_job_outputs(job, automatic_name)
            job = update_job(job_id, packet_name=automatic_name)
        vision_errors = [
            note
            for page in report.pages
            for note in page.notes
            if note.startswith("Vision OCR error:")
        ]
        n_vision = sum(1 for page in report.pages if page.ocr_backend_used == "vision")
        summary = summarize_report(report)
        summary = merge_issue_reviews(summary, job.get("issue_reviews"))
        warnings = []
        if provider in {"anthropic", "openai"} and vision_errors:
            warnings.append(
                "Vision OCR failed on one or more pages. The report was generated "
                "from Tesseract/printed-text OCR where possible. First vision error: "
                + vision_errors[0]
            )
        if provider in {"anthropic", "openai"} and n_vision == 0:
            warnings.append(
                f"{provider} vision OCR did not process any pages. This is a partial "
                "Tesseract-only report and may miss handwriting."
            )
        summary["warnings"] = warnings
        final_job = update_job(
            job_id,
            status="complete",
            workflow_status="Manual Review Pending",
            completed_at=utc_now(),
            message="Verification complete with OCR warnings" if warnings else "Verification complete",
            progress_percent=100,
            progress_stage="Verification complete",
            processed_pages=len(report.pages),
            total_pages=len(report.pages),
            summary=summary,
        )
        refresh_reviewed_output_files(final_job)
        add_notification(
            "Packet verification complete",
            f"{final_job.get('packet_name')} finished with {summary.get('fail_count', 0)} flag(s).",
            recipient=final_job.get("assignee", ""),
            url=f"/jobs/{job_id}",
            kind="warning" if summary.get("fail_count") else "success",
        )
    except Exception as exc:  # noqa: BLE001 - show useful operator error on job page
        failed_job = update_job(
            job_id,
            status="failed",
            completed_at=utc_now(),
            message=str(exc),
            progress_stage="Verification failed",
            error_trace=traceback.format_exc(),
        )
        add_notification(
            "Packet verification failed",
            f"{failed_job.get('packet_name')}: {exc}",
            recipient=failed_job.get("assignee", ""),
            url=f"/jobs/{job_id}",
            kind="error",
        )
    finally:
        packet_run_semaphore.release()


@app.get("/setup", response_class=HTMLResponse)
async def setup_form(request: Request) -> HTMLResponse:
    if load_users():
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        "setup.html",
        {"request": request, "app_name": APP_NAME, "error": ""},
    )


@app.post("/setup")
async def setup_admin(
    email: str = Form(...),
    password: str = Form(...),
    name: str = Form(""),
) -> RedirectResponse:
    if load_users():
        return RedirectResponse(url="/", status_code=303)
    create_user(email=email, password=password, role="Admin", name=name or "Admin")
    token = create_session(email)
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        SESSION_COOKIE,
        token,
        max_age=AUTH_SESSION_DAYS * 24 * 60 * 60,
        httponly=True,
        secure=AUTH_COOKIE_SECURE,
        samesite="lax",
    )
    return response


@app.get("/login", response_class=HTMLResponse)
async def login_form(request: Request, next: str = "/") -> HTMLResponse:
    if getattr(request.state, "user", None):
        return RedirectResponse(url=safe_next_path(next), status_code=303)
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "app_name": APP_NAME, "next": safe_next_path(next), "error": ""},
    )


@app.post("/login")
async def login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    next: str = Form("/"),
) -> Any:
    clean_email = normalize_email(email)
    user = load_users().get(clean_email)
    if not user or not user.get("active", True) or not verify_password(password, user.get("password_hash", "")):
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "app_name": APP_NAME,
                "next": safe_next_path(next),
                "error": "Email or password is incorrect.",
            },
            status_code=401,
        )
    token = create_session(clean_email)
    response = RedirectResponse(url=safe_next_path(next), status_code=303)
    response.set_cookie(
        SESSION_COOKIE,
        token,
        max_age=AUTH_SESSION_DAYS * 24 * 60 * 60,
        httponly=True,
        secure=AUTH_COOKIE_SECURE,
        samesite="lax",
    )
    return response


@app.get("/logout")
async def logout(request: Request) -> RedirectResponse:
    expire_session(request.cookies.get(SESSION_COOKIE, ""))
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/users", response_class=HTMLResponse)
async def users_page(request: Request) -> HTMLResponse:
    users = sorted(load_users().values(), key=lambda item: item.get("email", ""))
    safe_users = [{k: v for k, v in user.items() if k != "password_hash"} for user in users]
    return templates.TemplateResponse(
        "users.html",
        {"request": request, "app_name": APP_NAME, "users": safe_users, "roles": list(ROLE_ORDER)},
    )


@app.get("/notifications", response_class=HTMLResponse)
async def notifications_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "notifications.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "notifications": user_notifications(request.state.user),
        },
    )


@app.post("/notifications/read")
async def mark_notifications_read(request: Request) -> RedirectResponse:
    email = normalize_email(request.state.user.get("email", ""))
    items = load_json_file(NOTIFICATIONS_FILE, [])
    for item in items:
        if not item.get("recipient") or item.get("recipient") in {email, "admin"}:
            readers = set(item.get("read_by", []))
            readers.add(email)
            item["read_by"] = sorted(readers)
    save_json_file(NOTIFICATIONS_FILE, items)
    return RedirectResponse(url="/notifications", status_code=303)


@app.get("/learned-rules", response_class=HTMLResponse)
async def learned_rules_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "learned_rules.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "rules": list(reversed(load_json_file(LEARNED_RULES_FILE, []))),
        },
    )


@app.post("/learned-rules/{rule_id}")
async def review_learned_rule(
    request: Request,
    rule_id: str,
    status: str = Form(...),
) -> RedirectResponse:
    if status not in {"Approved", "Rejected", "Pending"}:
        raise HTTPException(status_code=400, detail="Unsupported rule status.")
    rules = load_json_file(LEARNED_RULES_FILE, [])
    rule = next((item for item in rules if item.get("id") == rule_id), None)
    if not rule:
        raise HTTPException(status_code=404, detail="Rule proposal not found.")
    rule.update({"status": status, "reviewed_by": actor_from_request(request), "reviewed_at": utc_now()})
    save_json_file(LEARNED_RULES_FILE, rules)
    return RedirectResponse(url="/learned-rules", status_code=303)


@app.get("/backups", response_class=HTMLResponse)
async def backups_page(request: Request) -> HTMLResponse:
    items = sorted(BACKUPS_DIR.glob("*.zip"), key=lambda path: path.stat().st_mtime, reverse=True)
    return templates.TemplateResponse(
        "backups.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "backups": [{"name": path.name, "size": path.stat().st_size} for path in items],
            "external_enabled": bool(os.environ.get("BACKUP_WEBHOOK_URL")),
        },
    )


@app.post("/backups")
async def create_backup() -> RedirectResponse:
    create_backup_archive()
    return RedirectResponse(url="/backups", status_code=303)


@app.get("/backups/{filename}")
async def download_backup(filename: str) -> FileResponse:
    path = (BACKUPS_DIR / filename).resolve()
    if BACKUPS_DIR.resolve() not in path.parents or not path.exists() or path.suffix != ".zip":
        raise HTTPException(status_code=404, detail="Backup not found.")
    return FileResponse(path, filename=path.name)


@app.post("/users")
async def add_user(
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("Reviewer"),
    name: str = Form(""),
) -> RedirectResponse:
    create_user(email=email, password=password, role=role, name=name)
    return RedirectResponse(url="/users", status_code=303)


@app.post("/users/update")
async def update_user(
    email: str = Form(...),
    role: str = Form("Viewer"),
    active: str = Form("true"),
    name: str = Form(""),
    password: str = Form(""),
) -> RedirectResponse:
    clean_email = normalize_email(email)
    users = load_users()
    user = users.get(clean_email)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if role not in ROLE_ORDER:
        raise HTTPException(status_code=400, detail="Unsupported role.")
    if password.strip() and len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters.")
    next_active = active == "true"
    if user.get("role") == "Admin" and user.get("active", True) and (role != "Admin" or not next_active):
        if active_admin_count(users) <= 1:
            raise HTTPException(status_code=400, detail="At least one active Admin user is required.")
    user["role"] = role
    user["active"] = next_active
    user["name"] = name.strip() or clean_email
    user["updated_at"] = utc_now()
    if password.strip():
        user["password_hash"] = hash_password(password)
    users[clean_email] = user
    save_users(users)
    return RedirectResponse(url="/users", status_code=303)


@app.get("/", response_class=HTMLResponse)
async def index(
    request: Request,
    q: str = "",
    workflow: str = "",
    assignee: str = "",
) -> HTMLResponse:
    jobs = sorted([public_job(j) for j in load_jobs().values()], key=lambda j: j.get("created_at", ""), reverse=True)
    query = q.strip().lower()
    if query:
        jobs = [
            job for job in jobs
            if query in " ".join([
                str(job.get("packet_name", "")),
                str(job.get("original_filename", "")),
                str((job.get("summary") or {}).get("sub_packets", "")),
                str(job.get("assignee", "")),
            ]).lower()
        ]
    if workflow:
        jobs = [job for job in jobs if (job.get("workflow_status") or "AI Processing") == workflow]
    if assignee:
        jobs = [job for job in jobs if (job.get("assignee") or "") == assignee]
    users = [
        {k: v for k, v in item.items() if k != "password_hash"}
        for item in load_users().values()
        if item.get("active", True) and item.get("role") in {"Admin", "Reviewer"}
    ]
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "jobs": jobs,
            "app_name": APP_NAME,
            "default_provider": DEFAULT_VISION_PROVIDER,
            "providers": PROVIDERS,
            "max_upload_mb": MAX_UPLOAD_MB,
            "query": q,
            "workflow": workflow,
            "assignee": assignee,
            "reviewers": sorted(users, key=lambda item: item.get("email", "")),
            "workflow_counts": packet_workflow_counts(jobs),
        },
    )


@app.get("/settings", response_class=HTMLResponse)
async def settings(request: Request, selected: str = "rules") -> HTMLResponse:
    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            "app_name": APP_NAME,
            **load_config_panels(selected=selected),
        },
    )


@app.get("/issues", response_class=HTMLResponse)
async def issue_board(request: Request) -> HTMLResponse:
    items = issue_board_items()
    return templates.TemplateResponse(
        "issues.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "issues": items,
            "counts": issue_counts(items),
        },
    )


@app.post("/issues/update")
async def update_issue_board(
    request: Request,
    board_key: str = Form(...),
    status: str = Form("Open"),
    assignee: str = Form(""),
    due_date: str = Form(""),
    priority: str = Form("Normal"),
    comment: str = Form(""),
) -> RedirectResponse:
    allowed = {"Open", "Corrected", "Accepted as Is", "False Positive", "Resolved"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported issue status.")
    board = load_json_file(BOARD_FILE, {})
    current = board.get(board_key, {})
    history = list(current.get("history") or [])
    history.append({
        "at": utc_now(),
        "status": status,
        "assignee": assignee.strip(),
        "due_date": due_date.strip(),
        "priority": priority.strip(),
        "comment": comment.strip(),
        "user": actor_from_request(request),
    })
    board[board_key] = {
        "status": status,
        "assignee": assignee.strip(),
        "due_date": due_date.strip(),
        "priority": priority.strip() or "Normal",
        "comment": comment.strip(),
        "updated_at": utc_now(),
        "history": history[-25:],
    }
    save_json_file(BOARD_FILE, board)
    return RedirectResponse(url="/issues", status_code=303)


@app.get("/inspections", response_class=HTMLResponse)
async def inspections(request: Request) -> HTMLResponse:
    items = sorted(load_json_file(INSPECTIONS_FILE, []), key=lambda x: x.get("created_at", ""), reverse=True)
    return templates.TemplateResponse(
        "inspections.html",
        {"request": request, "app_name": APP_NAME, "inspections": items},
    )


@app.post("/inspections")
async def create_inspection(
    request: Request,
    inspection_type: str = Form(...),
    inspector: str = Form(""),
    area: str = Form(""),
    date: str = Form(""),
    result: str = Form("Pass"),
    observations: str = Form(""),
    corrective_action: str = Form(""),
) -> RedirectResponse:
    items = load_json_file(INSPECTIONS_FILE, [])
    items.append({
        "id": uuid4().hex[:10],
        "created_at": utc_now(),
        "created_by": actor_from_request(request),
        "inspection_type": inspection_type.strip(),
        "inspector": inspector.strip(),
        "area": area.strip(),
        "date": date.strip(),
        "result": result,
        "observations": observations.strip(),
        "corrective_action": corrective_action.strip(),
    })
    save_json_file(INSPECTIONS_FILE, items)
    return RedirectResponse(url="/inspections", status_code=303)


@app.get("/forms", response_class=HTMLResponse)
async def food_safety_forms(request: Request, q: str = "", status: str = "") -> HTMLResponse:
    jobs = sorted(load_form_jobs().values(), key=lambda item: item.get("created_at", ""), reverse=True)
    query = q.strip().lower()
    if query:
        jobs = [
            item for item in jobs
            if query in " ".join([
                str(item.get("name", "")),
                str(item.get("original_filename", "")),
                str(item.get("template_label", "")),
                str(item.get("period", "")),
            ]).lower()
        ]
    if status:
        jobs = [item for item in jobs if (item.get("workflow_status") or item.get("status")) == status]
    form_templates = [
        item for item in get_form_templates().values()
        if item.get("active", True) and item.get("category") == "Food Safety"
    ]
    return templates.TemplateResponse(
        "forms.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "jobs": jobs,
            "templates": form_templates,
            "query": q,
            "selected_status": status,
            "calendar_month": request.query_params.get("month") or datetime.now().strftime("%Y-%m"),
            "calendar": production_calendar(),
            "missing_forms": missing_form_expectations(
                request.query_params.get("month") or datetime.now().strftime("%Y-%m")
            ),
        },
    )


@app.post("/forms/calendar")
async def update_production_calendar(
    date: str = Form(...),
    production_status: str = Form("non-production"),
    note: str = Form(""),
) -> RedirectResponse:
    try:
        parsed = datetime.strptime(date, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Enter a valid calendar date.") from exc
    calendar = production_calendar()
    closed = calendar.setdefault("non_production_dates", {})
    notes = calendar.setdefault("notes", {})
    if production_status == "production":
        closed.pop(date, None)
    else:
        closed[date] = note.strip() or "Non-production day"
    if note.strip():
        notes[date] = note.strip()
    save_json_file(PRODUCTION_CALENDAR_FILE, calendar)
    return RedirectResponse(url=f"/forms?month={parsed.strftime('%Y-%m')}", status_code=303)


@app.post("/forms")
async def create_food_safety_form(
    background_tasks: BackgroundTasks,
    request: Request,
    pdf: UploadFile = File(...),
    template_code: str = Form(...),
    name: str = Form(""),
    period: str = Form(""),
    form_date: str = Form(""),
    department: str = Form(""),
) -> RedirectResponse:
    if not pdf.filename or not pdf.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Upload a PDF form.")
    template = get_form_templates().get(template_code)
    if not template or template.get("category") != "Food Safety":
        raise HTTPException(status_code=400, detail="Select a Food Safety form template.")
    form_job_id = uuid4().hex[:12]
    input_path = FORM_UPLOAD_DIR / f"{form_job_id}_{slugify(pdf.filename)}.pdf"
    output_dir = FORM_OUTPUT_DIR / form_job_id
    output_dir.mkdir(parents=True, exist_ok=True)
    size = 0
    with input_path.open("wb") as handle:
        while chunk := await pdf.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD_MB * 1024 * 1024:
                input_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=f"PDF exceeds {MAX_UPLOAD_MB} MB.")
            handle.write(chunk)
    job = {
        "id": form_job_id,
        "name": name.strip() or slugify(pdf.filename),
        "original_filename": pdf.filename,
        "template_code": template_code,
        "template_label": template.get("label"),
        "period": period.strip(),
        "form_date": form_date.strip(),
        "department": department.strip(),
        "input_path": str(input_path),
        "output_dir": str(output_dir),
        "status": "queued",
        "workflow_status": "Awaiting Verification",
        "message": "Queued for form verification",
        "created_by": actor_from_request(request),
        "created_at": utc_now(),
        "updated_at": utc_now(),
    }
    update_form_job(form_job_id, **job)
    background_tasks.add_task(run_form_verification, form_job_id)
    return RedirectResponse(url=f"/forms/{form_job_id}", status_code=303)


@app.get("/forms/{form_job_id}", response_class=HTMLResponse)
async def food_safety_form_detail(request: Request, form_job_id: str) -> HTMLResponse:
    job = get_form_job(form_job_id)
    return templates.TemplateResponse(
        "form_job.html",
        {"request": request, "app_name": APP_NAME, "job": job},
    )


@app.post("/forms/{form_job_id}/workflow")
async def update_food_safety_form_workflow(
    request: Request,
    form_job_id: str,
    workflow_status: str = Form(...),
    reviewer_comment: str = Form(""),
) -> RedirectResponse:
    allowed = {"Awaiting Scan", "Awaiting Verification", "Needs Correction", "Approved", "Filed"}
    if workflow_status not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported form workflow status.")
    job = get_form_job(form_job_id)
    history = list(job.get("history") or [])
    history.append({
        "at": utc_now(),
        "user": actor_from_request(request),
        "status": workflow_status,
        "comment": reviewer_comment.strip(),
    })
    update_form_job(
        form_job_id,
        workflow_status=workflow_status,
        reviewer_comment=reviewer_comment.strip(),
        history=history[-50:],
    )
    if workflow_status in {"Approved", "Filed"}:
        auto_file_form(get_form_job(form_job_id))
    add_notification(
        "Food Safety workflow updated",
        f"{job.get('name')} was moved to {workflow_status}.",
        recipient=job.get("created_by", ""),
        url=f"/forms/{form_job_id}",
    )
    return RedirectResponse(url=f"/forms/{form_job_id}", status_code=303)


@app.get("/forms/{form_job_id}/file")
async def food_safety_form_file(form_job_id: str) -> FileResponse:
    job = get_form_job(form_job_id)
    path = Path(job["input_path"]).resolve()
    if FORM_UPLOAD_DIR.resolve() not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="Form file not found")
    return FileResponse(path, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{path.name}"'})


@app.get("/templates", response_class=HTMLResponse)
async def form_templates(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "templates.html",
        {"request": request, "app_name": APP_NAME, "templates": get_form_templates().values()},
    )


@app.post("/templates")
async def save_template(
    code: str = Form(...),
    label: str = Form(""),
    fields: str = Form(""),
    active: str = Form("true"),
    category: str = Form("Packet"),
    required_terms: str = Form(""),
    require_date: str = Form("false"),
    required_signatures: int = Form(0),
    frequency: str = Form("As needed"),
    tracking_enabled: str = Form("false"),
    related_templates: str = Form(""),
    regions: str = Form("[]"),
    sample_pdf: Optional[UploadFile] = File(None),
) -> RedirectResponse:
    sample_filename = ""
    clean_code = re.sub(r"[^A-Za-z0-9_]+", "_", code.strip().upper()).strip("_")
    if sample_pdf and sample_pdf.filename:
        if not sample_pdf.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Template sample must be a PDF.")
        sample_path = TEMPLATE_SAMPLE_DIR / f"{clean_code}.pdf"
        with sample_path.open("wb") as handle:
            while chunk := await sample_pdf.read(1024 * 1024):
                handle.write(chunk)
        sample_filename = sample_path.name
    save_form_template(
        code, label, fields, active == "true",
        category=category,
        required_terms_text=required_terms,
        require_date=require_date == "true",
        required_signatures=required_signatures,
        regions_text=regions,
        sample_filename=sample_filename,
        frequency=frequency,
        tracking_enabled=tracking_enabled == "true",
        related_templates_text=related_templates,
    )
    return RedirectResponse(url="/templates", status_code=303)


@app.get("/templates/sample/{code}")
async def template_sample(code: str) -> FileResponse:
    item = get_form_templates().get(code)
    if not item or not item.get("sample_filename"):
        raise HTTPException(status_code=404, detail="Template sample not found")
    path = (TEMPLATE_SAMPLE_DIR / item["sample_filename"]).resolve()
    if TEMPLATE_SAMPLE_DIR.resolve() not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="Template sample not found")
    return FileResponse(path, media_type="application/pdf")


@app.get("/templates/sample/{code}/page.png")
async def template_sample_page(code: str) -> FileResponse:
    item = get_form_templates().get(code)
    if not item or not item.get("sample_filename"):
        raise HTTPException(status_code=404, detail="Template sample not found")
    pdf_path = (TEMPLATE_SAMPLE_DIR / item["sample_filename"]).resolve()
    if TEMPLATE_SAMPLE_DIR.resolve() not in pdf_path.parents or not pdf_path.exists():
        raise HTTPException(status_code=404, detail="Template sample not found")
    image_path = render_pdf_page_to_png(pdf_path, 0, TEMPLATE_SAMPLE_DIR / f"{code}.png")
    return FileResponse(image_path, media_type="image/png")


@app.get("/audit", response_class=HTMLResponse)
async def audit_prep(request: Request) -> HTMLResponse:
    jobs = sorted([public_job(j) for j in load_jobs().values()], key=lambda j: j.get("created_at", ""), reverse=True)
    inspections_data = sorted(load_json_file(INSPECTIONS_FILE, []), key=lambda x: x.get("created_at", ""), reverse=True)
    return templates.TemplateResponse(
        "audit.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "jobs": jobs[:100],
            "inspections": inspections_data[:100],
            "open_issues": issue_counts(issue_board_items())["open"],
        },
    )


@app.post("/audit/generate")
async def generate_audit(month: str = Form("")) -> RedirectResponse:
    month_key = month.strip() or datetime.now(timezone.utc).strftime("%Y-%m")
    jobs = [public_job(j) for j in load_jobs().values() if str(j.get("created_at", "")).startswith(month_key)]
    inspections_data = [i for i in load_json_file(INSPECTIONS_FILE, []) if str(i.get("date") or i.get("created_at", "")).startswith(month_key)]
    out = MONTHLY_DIR / f"california-fruit-audit-prep-{slugify(month_key)}.xlsx"
    write_audit_workbook(out, month_key, jobs, inspections_data)
    return RedirectResponse(url=f"/audit/download/{out.name}", status_code=303)


@app.get("/audit/download/{filename}")
async def download_audit(filename: str) -> FileResponse:
    path = (MONTHLY_DIR / filename).resolve()
    if MONTHLY_DIR.resolve() not in path.parents or not path.exists():
        raise HTTPException(status_code=404, detail="Audit file not found")
    return FileResponse(path, filename=path.name)


@app.post("/settings/{config_name}", response_class=HTMLResponse)
async def save_settings(
    request: Request,
    config_name: str,
    content: str = Form(...),
) -> HTMLResponse:
    path = config_file_path(config_name)
    try:
        yaml.safe_load(content) or {}
    except yaml.YAMLError as exc:
        return templates.TemplateResponse(
            "settings.html",
            {
                "request": request,
                "app_name": APP_NAME,
                **load_config_panels(
                    selected=config_name,
                    error=f"{path.name} was not saved because the YAML is invalid: {exc}",
                ),
            },
            status_code=400,
        )
    backup = path.with_suffix(path.suffix + f".{utc_now().replace(':', '-')}.bak")
    shutil.copy2(path, backup)
    path.write_text(content, encoding="utf-8")
    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            "app_name": APP_NAME,
            **load_config_panels(selected=config_name, message=f"Saved {path.name}; backup created as {backup.name}."),
        },
    )


@app.post("/jobs")
async def create_job(
    background_tasks: BackgroundTasks,
    pdf: UploadFile = File(...),
    packet_name: str = Form(""),
    vision_provider: str = Form(DEFAULT_VISION_PROVIDER),
) -> RedirectResponse:
    if not pdf.filename or not pdf.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Upload a PDF packet.")
    provider = vision_provider.strip().lower()
    if provider not in {"mock", "anthropic", "openai"}:
        raise HTTPException(status_code=400, detail="Unsupported vision provider.")

    job_id = uuid4().hex[:12]
    safe_name = slugify(packet_name or pdf.filename)
    input_path = UPLOAD_DIR / f"{job_id}_{slugify(pdf.filename)}.pdf"
    output_dir = OUTPUT_DIR / job_id
    output_dir.mkdir(parents=True, exist_ok=True)

    size = 0
    with input_path.open("wb") as out:
        while chunk := await pdf.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD_MB * 1024 * 1024:
                input_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=f"PDF exceeds {MAX_UPLOAD_MB} MB.")
            out.write(chunk)

    job = {
        "id": job_id,
        "packet_name": safe_name,
        "packet_name_user_supplied": bool(packet_name.strip()),
        "original_filename": pdf.filename,
        "input_path": str(input_path),
        "output_dir": str(output_dir),
        "vision_provider": provider,
        "status": "queued",
        "workflow_status": "AI Processing",
        "assignee": "",
        "due_date": "",
        "message": "Queued for verification",
        "progress_percent": 1,
        "progress_stage": "Queued for verification",
        "processed_pages": 0,
        "total_pages": 0,
        "created_at": utc_now(),
        "updated_at": utc_now(),
        "summary": None,
    }
    update_job(job_id, **job)
    background_tasks.add_task(run_verification, job_id)
    return RedirectResponse(url=f"/jobs/{job_id}", status_code=303)


@app.get("/jobs/{job_id}", response_class=HTMLResponse)
async def job_detail(request: Request, job_id: str) -> HTMLResponse:
    raw_job = get_job(job_id)
    refresh_reviewed_output_files(raw_job)
    job = public_job(raw_job)
    return templates.TemplateResponse(
        "job.html",
        {
            "request": request,
            "job": job,
            "files": output_files(job),
            "app_name": APP_NAME,
            "pages": page_catalog(job),
            "reviewers": sorted(
                [
                    {k: v for k, v in item.items() if k != "password_hash"}
                    for item in load_users().values()
                    if item.get("active", True) and item.get("role") in {"Admin", "Reviewer"}
                ],
                key=lambda item: item.get("email", ""),
            ),
        },
    )


@app.post("/jobs/{job_id}/workflow")
async def update_job_workflow(
    request: Request,
    job_id: str,
    workflow_status: str = Form("Manual Review Pending"),
    assignee: str = Form(""),
    due_date: str = Form(""),
) -> RedirectResponse:
    allowed = {"AI Processing", "Manual Review Pending", "Correction Required", "Approved", "Filed"}
    if workflow_status not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported workflow status.")
    job = get_job(job_id)
    audit_event(
        job,
        "workflow_updated",
        f"Workflow set to {workflow_status}",
        user=actor_from_request(request),
        assignee=assignee.strip(),
        due_date=due_date.strip(),
    )
    update_job(
        job_id,
        workflow_status=workflow_status,
        assignee=assignee.strip(),
        due_date=due_date.strip(),
        audit_trail=job.get("audit_trail"),
    )
    if assignee.strip():
        add_notification(
            "Packet assigned",
            f"{job.get('packet_name')} was assigned to you with status {workflow_status}.",
            recipient=assignee.strip(),
            url=f"/jobs/{job_id}",
        )
    return RedirectResponse(url=f"/jobs/{job_id}", status_code=303)


@app.get("/jobs/{job_id}/fields", response_class=HTMLResponse)
async def job_fields(request: Request, job_id: str) -> HTMLResponse:
    job = public_job(get_job(job_id))
    return templates.TemplateResponse(
        "fields.html",
        {
            "request": request,
            "app_name": APP_NAME,
            "job": job,
            "fields": field_confidence_rows(job),
        },
    )


@app.get("/api/jobs/{job_id}")
async def job_status(job_id: str) -> Dict[str, Any]:
    job = public_job(get_job(job_id))
    return {**job, "files": output_files(job)}


@app.post("/jobs/{job_id}/issues/{issue_id}")
async def update_issue_review(
    request: Request,
    job_id: str,
    issue_id: str,
    status: str = Form(...),
    comment: str = Form(""),
    propose_rule: str = Form(""),
) -> RedirectResponse:
    allowed = {"Open", "Corrected", "Accepted as Is", "False Positive", "Resolved"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported issue status.")
    with jobs_lock:
        jobs = load_jobs()
        job = jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        summary = job.get("summary") or {}
        issue = next((item for item in summary.get("issues", []) if item.get("id") == issue_id), None)
        if not issue:
            raise HTTPException(status_code=404, detail="Issue not found")
        reviews = dict(job.get("issue_reviews") or {})
        reviews[issue["key"]] = {
            "status": status,
            "comment": comment.strip(),
            "updated_at": utc_now(),
        }
        job["issue_reviews"] = reviews
        issue.update(reviews[issue["key"]])
        audit_event(
            job,
            "issue_review",
            f"{issue_id} marked {status}",
            issue_id=issue_id,
            issue_type=issue.get("issue_type"),
            pages=issue.get("pages", []),
            comment=comment.strip(),
            user=actor_from_request(request),
        )
        job["updated_at"] = utc_now()
        jobs[job_id] = job
        save_jobs(jobs)
    if status == "False Positive" and propose_rule == "true":
        form_code = ""
        pages = issue.get("pages") or []
        catalog = {item["page_no"]: item for item in page_catalog(job)}
        if pages:
            form_code = str(catalog.get(int(pages[0]), {}).get("form_code") or "")
        rules = load_json_file(LEARNED_RULES_FILE, [])
        normalized_name = re.sub(r"\(p\d+\)", "(p#)", issue.get("name", "").lower())
        proposal = {
            "id": f"FPR-{uuid4().hex[:8].upper()}",
            "status": "Pending",
            "name_pattern": normalized_name,
            "detail_pattern": "",
            "form_code": form_code,
            "example_name": issue.get("name"),
            "example_detail": issue.get("detail"),
            "source_job_id": job_id,
            "created_by": actor_from_request(request),
            "created_at": utc_now(),
        }
        rules.append(proposal)
        save_json_file(LEARNED_RULES_FILE, rules)
        add_notification(
            "False-positive rule awaiting approval",
            f"{proposal['id']} was proposed from {job.get('packet_name')}.",
            recipient="admin",
            url="/learned-rules",
            kind="warning",
        )
    refresh_reviewed_output_files(get_job(job_id))
    return RedirectResponse(url=f"/jobs/{job_id}#review-queue", status_code=303)


@app.post("/jobs/{job_id}/delete")
async def delete_job(request: Request, job_id: str) -> RedirectResponse:
    job = get_job(job_id)
    upload_root = UPLOAD_DIR.resolve()
    output_root = OUTPUT_DIR.resolve()
    input_path = Path(job["input_path"]).resolve()
    output_path = Path(job["output_dir"]).resolve()
    if upload_root not in input_path.parents or output_root not in output_path.parents:
        raise HTTPException(status_code=400, detail="Packet paths are outside the managed data directory.")
    if input_path.exists():
        input_path.unlink()
    if output_path.exists():
        shutil.rmtree(output_path)
    jobs = load_jobs()
    jobs.pop(job_id, None)
    save_jobs(jobs)
    board = load_json_file(BOARD_FILE, {})
    board = {key: value for key, value in board.items() if not key.startswith(f"{job_id}:")}
    save_json_file(BOARD_FILE, board)
    add_notification(
        "Packet deleted",
        f"{job.get('packet_name')} was deleted by {actor_from_request(request)}.",
        recipient="admin",
        kind="warning",
    )
    return RedirectResponse(url="/", status_code=303)


@app.post("/jobs/{job_id}/signoff")
async def save_signoff(
    request: Request,
    job_id: str,
    reviewer_name: str = Form(...),
    decision: str = Form(...),
    comments: str = Form(""),
) -> RedirectResponse:
    allowed = {"Approved", "Approved with exceptions", "Rejected", "Needs correction"}
    if decision not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported sign-off decision.")
    with jobs_lock:
        jobs = load_jobs()
        job = jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        signoff = {
            "reviewer_name": reviewer_name.strip(),
            "decision": decision,
            "comments": comments.strip(),
            "signed_at": utc_now(),
        }
        job["review_signoff"] = signoff
        audit_event(
            job,
            "review_signoff",
            f"{reviewer_name.strip()} signed: {decision}",
            decision=decision,
            comments=comments.strip(),
            user=actor_from_request(request),
        )
        job["updated_at"] = utc_now()
        jobs[job_id] = job
        save_jobs(jobs)
    return RedirectResponse(url=f"/jobs/{job_id}#review-signoff", status_code=303)


@app.get("/jobs/{job_id}/download/{filename}")
async def download(job_id: str, filename: str) -> FileResponse:
    job = get_job(job_id)
    path = output_file_path(job, filename)
    return FileResponse(path, filename=filename)


@app.get("/jobs/{job_id}/view/{filename}")
async def view_file(job_id: str, filename: str) -> FileResponse:
    job = get_job(job_id)
    path = output_file_path(job, filename)
    media_types = {
        ".pdf": "application/pdf",
        ".png": "image/png",
        ".csv": "text/csv; charset=utf-8",
        ".json": "application/json; charset=utf-8",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return FileResponse(
        path,
        media_type=media_types.get(path.suffix.lower(), "application/octet-stream"),
        headers={"Content-Disposition": f'inline; filename="{path.name}"'},
    )


@app.get("/jobs/{job_id}/preview/{filename}")
async def preview_file(job_id: str, filename: str) -> JSONResponse:
    job = get_job(job_id)
    path = output_file_path(job, filename)
    return JSONResponse(preview_table_file(path))


@app.get("/jobs/{job_id}/page/{page_no}.png")
async def page_image(job_id: str, page_no: int) -> FileResponse:
    job = get_job(job_id)
    path = page_image_path(job, page_no)
    return FileResponse(path, media_type="image/png")


@app.get("/jobs/{job_id}/page/{page_no}.pdf")
async def page_pdf(job_id: str, page_no: int) -> FileResponse:
    if page_no < 1:
        raise HTTPException(status_code=404, detail="Page not found")
    from pypdf import PdfReader, PdfWriter

    job = get_job(job_id)
    source = PdfReader(job["input_path"])
    if page_no > len(source.pages):
        raise HTTPException(status_code=404, detail="Page not found")
    page_dir = Path(job["output_dir"]) / "page_exports"
    page_dir.mkdir(parents=True, exist_ok=True)
    path = page_dir / f"{job['packet_name']}_page_{page_no:03d}.pdf"
    if not path.exists():
        writer = PdfWriter()
        writer.add_page(source.pages[page_no - 1])
        with path.open("wb") as handle:
            writer.write(handle)
    return FileResponse(path, media_type="application/pdf", filename=path.name)


@app.get("/jobs/{job_id}/compare/{event_id}/{side}.png")
async def compare_page(job_id: str, event_id: str, side: str) -> FileResponse:
    if side not in {"before", "after"}:
        raise HTTPException(status_code=404, detail="Comparison side not found")
    job = get_job(job_id)
    event = next((item for item in job.get("audit_trail") or [] if item.get("id") == event_id), None)
    if not event or event.get("action") != "page_replaced":
        raise HTTPException(status_code=404, detail="Replacement event not found")
    page_no = int(event.get("page_no") or 1)
    out_dir = Path(job["output_dir"]).resolve()
    compare_dir = out_dir / "versions" / "compare" / event_id
    if side == "before":
        pdf_path = Path(event.get("backup_pdf", "")).resolve()
        page_index = page_no - 1
    else:
        pdf_path = Path(event.get("replacement_pdf", "")).resolve()
        page_index = 0
    versions_dir = (out_dir / "versions").resolve()
    if versions_dir not in pdf_path.parents or not pdf_path.exists():
        raise HTTPException(status_code=404, detail="Comparison PDF not found")
    image_path = render_pdf_page_to_png(pdf_path, page_index, compare_dir / f"{side}.png")
    return FileResponse(image_path, media_type="image/png")


@app.post("/jobs/{job_id}/rerun")
async def rerun_job(background_tasks: BackgroundTasks, job_id: str) -> RedirectResponse:
    job = get_job(job_id)
    out_dir = Path(job["output_dir"])
    if out_dir.exists():
        shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
    update_job(
        job_id,
        status="queued",
        message="Queued for re-run",
        progress_percent=1,
        progress_stage="Queued for re-run",
        processed_pages=0,
        total_pages=0,
        summary=None,
    )
    background_tasks.add_task(run_verification, job_id)
    return RedirectResponse(url=f"/jobs/{job_id}", status_code=303)


@app.post("/jobs/{job_id}/replace-page")
async def replace_page(
    request: Request,
    background_tasks: BackgroundTasks,
    job_id: str,
    page_no: int = Form(...),
    replacement_pdf: UploadFile = File(...),
) -> RedirectResponse:
    if page_no < 1:
        raise HTTPException(status_code=400, detail="Page number must be 1 or greater.")
    if not replacement_pdf.filename or not replacement_pdf.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Upload a PDF replacement page.")
    job = get_job(job_id)
    out_dir = Path(job["output_dir"])
    versions_dir = out_dir / "versions"
    versions_dir.mkdir(parents=True, exist_ok=True)
    stamp = utc_now().replace(":", "-")
    replacement_path = versions_dir / f"{stamp}_replacement_p{page_no:02d}_{slugify(replacement_pdf.filename)}.pdf"

    size = 0
    with replacement_path.open("wb") as out:
        while chunk := await replacement_pdf.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD_MB * 1024 * 1024:
                replacement_path.unlink(missing_ok=True)
                raise HTTPException(status_code=413, detail=f"Replacement PDF exceeds {MAX_UPLOAD_MB} MB.")
            out.write(chunk)

    try:
        from pypdf import PdfReader, PdfWriter

        source = PdfReader(job["input_path"])
        replacement = PdfReader(str(replacement_path))
        if page_no > len(source.pages):
            raise HTTPException(status_code=400, detail=f"Packet only has {len(source.pages)} pages.")
        if len(replacement.pages) < 1:
            raise HTTPException(status_code=400, detail="Replacement PDF has no pages.")
        previous_input = Path(job["input_path"])
        backup_input = versions_dir / f"{stamp}_before_replace_p{page_no:02d}_{previous_input.name}"
        shutil.copy2(previous_input, backup_input)
        writer = PdfWriter()
        for i, page in enumerate(source.pages, start=1):
            writer.add_page(replacement.pages[0] if i == page_no else page)
        new_input = UPLOAD_DIR / f"{job_id}_{job['packet_name']}_page-{page_no:02d}-replaced.pdf"
        with new_input.open("wb") as f:
            writer.write(f)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not replace page: {exc}") from exc

    for child in list(out_dir.iterdir()):
        if child.name == "versions":
            continue
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink(missing_ok=True)

    job = update_job(
        job_id,
        input_path=str(new_input),
        status="queued",
        message=f"Page {page_no} replaced; queued for re-run",
        progress_percent=1,
        progress_stage="Queued after page replacement",
        processed_pages=0,
        total_pages=0,
        summary=None,
    )
    audit_event(
        job,
        "page_replaced",
        f"Page {page_no} replaced and queued for verification",
        page_no=page_no,
        backup_pdf=str(backup_input),
        replacement_pdf=str(replacement_path),
        new_packet_pdf=str(new_input),
        user=actor_from_request(request),
    )
    update_job(job_id, audit_trail=job.get("audit_trail"))
    background_tasks.add_task(run_verification, job_id)
    return RedirectResponse(url=f"/jobs/{job_id}", status_code=303)


@app.get("/healthz")
async def healthz() -> Dict[str, str]:
    return {"status": "ok", "app_version": APP_VERSION}


@app.get("/diagnostics")
async def diagnostics() -> Dict[str, Any]:
    css_path = ROOT / "app" / "static" / "app.css"
    return {
        "status": "ok",
        "app_version": APP_VERSION,
        "vision_provider": DEFAULT_VISION_PROVIDER,
        "anthropic_model": ANTHROPIC_MODEL,
        "openai_model": OPENAI_MODEL,
        "full_packet_field_discovery": FULL_PACKET_FIELD_DISCOVERY,
        "preview_max_rows": PREVIEW_MAX_ROWS,
        "preview_max_cols": PREVIEW_MAX_COLS,
        "anthropic_key_present": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "openai_key_present": bool(os.environ.get("OPENAI_API_KEY")),
        "tesseract_path": shutil.which("tesseract"),
        "pdftoppm_path": shutil.which("pdftoppm"),
        "css_exists": css_path.exists(),
        "css_size": css_path.stat().st_size if css_path.exists() else 0,
        "data_dir": str(DATA_DIR),
        "config_dir": str(CONFIG_DIR),
        "bundled_config_dir": str(BUNDLED_CONFIG_DIR),
    }


@app.get("/diagnostics/anthropic")
async def diagnostics_anthropic() -> Dict[str, Any]:
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not set")
    try:
        import anthropic

        client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
        msg = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=16,
            messages=[{"role": "user", "content": "Reply with exactly: ok"}],
        )
        return {
            "status": "ok",
            "model": ANTHROPIC_MODEL,
            "response": msg.content[0].text,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/diagnostics/openai")
async def diagnostics_openai() -> Dict[str, Any]:
    if not os.environ.get("OPENAI_API_KEY"):
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY is not set")
    try:
        payload = {
            "model": OPENAI_MODEL,
            "input": "Reply with exactly: ok",
            "reasoning": {"effort": os.environ.get("OPENAI_REASONING_EFFORT", "none")},
            "max_output_tokens": 16,
        }
        req = urllib.request.Request(
            "https://api.openai.com/v1/responses",
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {os.environ['OPENAI_API_KEY']}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        text = data.get("output_text") or ""
        if not text:
            chunks = []
            for item in data.get("output", []) or []:
                for content in item.get("content", []) or []:
                    if content.get("text"):
                        chunks.append(str(content["text"]))
            text = "\n".join(chunks)
        return {
            "status": "ok",
            "model": OPENAI_MODEL,
            "response": text.strip(),
        }
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(status_code=500, detail=f"OpenAI API error {exc.code}: {detail}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
